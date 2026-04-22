# =============================================================================
# app/api/v1/endpoints/defects.py  — PRODUCTION READY (Shore + Vessel)
#
# WHAT THIS FILE CONTAINS (every feature, nothing removed):
#   ✅ IST timezone — all date display/export uses IST
#   ✅ get_container_client() called as function — NOT module-level var
#   ✅ Live feed calls after every write (non-blocking, non-fatal)
#   ✅ GET  /                         — all defects with filters, ordered
#   ✅ GET  /export                   — Excel export, Sheet1 list + Sheet2 images
#   ✅ GET  /sas                      — upload SAS URL generation
#   ✅ GET  /{id}/validate-images     — check mandatory images before close
#   ✅ GET  /import-template          — Shore template (vessel dropdown from DB)
#   ✅ GET  /import-template-vessel   — Vessel template (no vessel column)
#   ✅ POST /import                   — Bulk import (openpyxl, "Area of Concern",
#                                       batch insert, SyncQueue, PR split)
#   ✅ GET  /{defect_id}/vessel-users — vessel crew + shore/admin for @mentions
#   ✅ POST /                         — create defect (via DefectService)
#   ✅ GET  /{defect_id}              — single defect
#   ✅ PATCH /{defect_id}             — update defect (status machine via service)
#   ✅ PATCH /{defect_id}/shore-close — shore direct closure (50 char remarks)
#   ✅ PATCH /{defect_id}/close       — legacy close with evidence
#   ✅ DELETE /{defect_id}            — soft delete
#   ✅ POST /threads                  — create thread (is_internal, @mention filter)
#   ✅ GET  /{defect_id}/threads      — get threads (role-based internal filter)
#   ✅ POST /attachments              — create attachment (1MB limit, idempotent)
#   ✅ GET  /{id}/threads/{tid}/attachments/{aid}/url — SAS download URL
#   ✅ POST /images                   — save image metadata (idempotent)
#   ✅ GET  /{defect_id}/images/{image_type} — get before/after images
#   ✅ DELETE /{defect_id}/images/{image_id} — delete image
#   ✅ POST /pr-entries               — create PR entry + live feed
#   ✅ PATCH /pr-entries/{pr_id}      — update PR entry
#   ✅ GET  /{defect_id}/pr-entries   — list active PR entries
#   ✅ DELETE /pr-entries/{pr_id}     — soft delete PR entry
#   ✅ create_system_thread()         — helper for system messages
#   ✅ COLUMN_MAP uses "Area of Concern" label (not "Equipment")
#   ✅ DefectService delegation for all DB writes + SyncQueue
#   ✅ _should_sync() imported and used for batch import SyncQueue writes
#   ✅ Batch insert in bulk import (1000-row chunks — much faster than ORM loop)
#   ✅ Email background tasks on every write
# =============================================================================

import uuid
from uuid import UUID
from datetime import datetime, timezone, timedelta
from app.core.config import settings
from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    BackgroundTasks,
    Query,
    UploadFile,
    File,
)
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from sqlalchemy import insert
import logging
import io
import xlsxwriter
import openpyxl
from PIL import Image as PILImage
from app.core.database_control import get_control_db
from app.core.database import get_db
from app.models.defect import Defect, Thread, Attachment, PrEntry, DefectImage
from app.models.user import User
from app.models.enums import UserRole, DefectStatus, DefectPriority, DefectSource
from app.models.vessel import Vessel
from app.models.sync import SyncQueue
from app.models.mariapps_pr_cache import MariappsPrCache
from app.schemas.defect import (
    DefectCreate,
    DefectUpdate,
    DefectResponse,
    ThreadCreate,
    ThreadResponse,
    AttachmentResponse,
    AttachmentBase,
    DefectCloseRequest,
    ShoreCloseRequest,
    VesselUserResponse,
    PrEntryCreate,
    PrEntryResponse,
    DefectImageCreate,
    DefectImageResponse,
    PrEntryUpdate,
)
from app.core.blob_storage import (
    generate_write_sas_url,
    generate_read_sas_url,
    get_container_client,
    download_blob_bytes,
)
import httpx
import base64
from app.api.deps import get_current_user
from app.services.email_service import send_defect_email
from app.services.notification_service import (
    notify_vessel_users,
    create_task_for_mentions,
)
from app.services.live_feed_service import (
    feed_defect_opened,
    feed_defect_closed,
    feed_priority_changed,
    feed_image_uploaded,
    feed_pic_mandatory_changed,
    feed_pr_added,
    feed_pr_invalid_format,
    feed_mention
)
from app.services.defect_service import DefectService, _should_sync
from app.models.defect import UserDefectFlag
from pydantic import BaseModel
from typing import Optional
logger = logging.getLogger(__name__)


class EmailDraftRequest(BaseModel):
    to_emails: Optional[list[str]] = None
    cc_emails: Optional[list[str]] = None
    subject: Optional[str] = None
    body_text: Optional[str] = None


def _build_email_content(defect, vessel_name: str, defect_ref: str) -> tuple[str, str]:
    """Return (subject, body_text) for a defect email draft."""
    date_identified = (
        defect.date_identified.strftime("%Y-%m-%d") if defect.date_identified else "N/A"
    )
    target_close = (
        defect.target_close_date.strftime("%Y-%m-%d") if defect.target_close_date else "N/A"
    )
    priority = (
        defect.priority.value if hasattr(defect.priority, "value") else str(defect.priority)
    )
    subject = f"Defect Raised - {defect_ref} | [{vessel_name}] {defect.title}"
    body_text = "\n".join(
        [
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "DEFECT REPORT — Maritime DRS",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"Vessel        : {vessel_name}",
            f"Title         : {defect.title}",
            f"Equipment     : {defect.equipment_name}",
            f"Priority      : {priority}",
            f"Identified On : {date_identified}",
            f"Target Close  : {target_close}",
            "",
            "Description:",
            defect.description,
            "",
            "━━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            "Generated from Workplace Platform",
        ]
    )
    return subject, body_text
router = APIRouter(redirect_slashes=False)

# IST timezone — used for all date display and Excel export
IST = timezone(timedelta(hours=5, minutes=30))

# Standard equipment list (shared by both import templates as "Area of Concern" dropdown)
EQUIPMENT_OPTIONS = [
    "HULL",
    "DECK",
    "SHIP ACCESS",
    "DECK MACHINERIES",
    "CARGO SYSTEM",
    "RADIO AND NAVIGATION",
    "BALLAST AND FUEL TANKS",
    "PAINT STORE WORKSHOP",
    "ACCOMMODATION SUPERSTRUCTURE",
    "ENGINE ROOM",
    "EMERGENCY MACHINERIES",
    "LIFE SAVING APPLIANCE",
    "FIRE FIGHTING APPLIANCE",
    "POLLUTION PREVENTION",
    "PMS",
    "ENERGY MANAGEMENT",
    "ELEVATOR",
    "MLC QHSE",
    "SECURITY",
    "CREW INTERACTION",
]


# =============================================================================
# HELPERS
# =============================================================================


def prepare_email_data(defect: Defect) -> dict:
    """Safely converts defect object to dictionary for email template."""
    return {
        "vessel_imo": defect.vessel_imo,
        "title": defect.title,
        "equipment_name": defect.equipment_name,
        "priority": (
            defect.priority.value
            if hasattr(defect.priority, "value")
            else str(defect.priority)
        ),
        "status": (
            defect.status.value
            if hasattr(defect.status, "value")
            else str(defect.status)
        ),
        "defect_source": (
            defect.defect_source.value
            if hasattr(defect.defect_source, "value")
            else str(defect.defect_source)
        ),
        "description": defect.description,
    }


# "Area of Concern" is the correct production column label (not "Equipment")
COLUMN_MAP = {
    "sno": ("S.No", lambda d, idx: idx),
    "defect_number": (
        "Defect ID",
        lambda d, idx: d.defect_number if d.defect_number else f"#{idx}",
    ),
    "date": (
        "Report Date",
        lambda d, idx: (
            d.date_identified.astimezone(IST).strftime("%Y-%m-%d")
            if d.date_identified
            else "-"
        ),
    ),
    "deadline": (
        "Deadline",
        lambda d, idx: (
            d.target_close_date.astimezone(IST).strftime("%Y-%m-%d")
            if d.target_close_date
            else "-"
        ),
    ),
    "source": (
        "Defect Source",
        lambda d, idx: (
            d.defect_source.value
            if hasattr(d.defect_source, "value")
            else str(d.defect_source)
        ),
    ),
    "equipment": (
        "Area of Concern",
        lambda d, idx: d.equipment_name,
    ),  # ← "Area of Concern" not "Equipment"
    "description": ("Description", lambda d, idx: d.description),
    "priority": (
        "Priority",
        lambda d, idx: (
            d.priority.value if hasattr(d.priority, "value") else str(d.priority)
        ),
    ),
    "status": (
        "Status",
        lambda d, idx: d.status.value if hasattr(d.status, "value") else str(d.status),
    ),
    "owner": ("Owner", lambda d, idx: "Owner" if d.is_owner else "Not Owner"),
    "dd": ("Dry Dock", lambda d, idx: "Yes" if d.is_dd else "No"),  # ✅ Added
    "pr_details": (
        "PR Number",
        lambda d, idx: ", ".join(
            [p.pr_number for p in d.pr_entries if not p.is_deleted]
        ),
    ),
    "closure_remarks": (
        "Closure Remarks",
        lambda d, idx: d.closure_remarks if d.closure_remarks else "-",
    ),
}


async def get_graph_token() -> str:
    """Get Microsoft Graph API token using client credentials."""
    url = f"https://login.microsoftonline.com/{settings.AZURE_TENANT_ID}/oauth2/v2.0/token"
    async with httpx.AsyncClient(verify=False) as client:
        response = await client.post(
            url,
            data={
                "grant_type": "client_credentials",
                "client_id": settings.AZURE_CLIENT_ID,
                "client_secret": settings.AZURE_CLIENT_SECRET,
                "scope": "https://graph.microsoft.com/.default",
            },
        )
        response.raise_for_status()
        return response.json()["access_token"]


async def create_outlook_draft(
    user_email: str,
    to_emails: list[str],
    subject: str,
    body_text: str,
    attachments: list[dict],  # [{"name": str, "content_type": str, "data": bytes}]
    cc_emails: list[str] = [],
) -> str:
    """
    Creates a draft email in user's Outlook via Graph API.
    Returns the web_link to open the draft directly.
    """
    token = await get_graph_token()

    # Build attachment list (base64 encoded)
    graph_attachments = []
    for att in attachments:
        graph_attachments.append(
            {
                "@odata.type": "#microsoft.graph.fileAttachment",
                "name": att["name"],
                "contentType": att.get("content_type", "application/octet-stream"),
                "contentBytes": base64.b64encode(att["data"]).decode("utf-8"),
            }
        )

    # Build email draft payload
    draft_payload = {
        "subject": subject,
        "importance": "Normal",
        "body": {"contentType": "Text", "content": body_text},
        "toRecipients": [{"emailAddress": {"address": email}} for email in to_emails],
        "ccRecipients": [{"emailAddress": {"address": email}} for email in cc_emails],
        "attachments": graph_attachments,
    }

    async with httpx.AsyncClient(verify=False) as client:
        response = await client.post(
            f"https://graph.microsoft.com/v1.0/users/{user_email}/messages",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json=draft_payload,
            timeout=30.0,
        )
        response.raise_for_status()
        data = response.json()
        message_id = data.get("id", "")
        logger.info(f"✅ Graph API Response: {data}")
        logger.info(f"✅ Draft message ID: {message_id}")
        logger.info(f"✅ Draft created for user: {user_email}")
        web_link = f"https://outlook.office365.com/mail/drafts"
        return web_link


@router.get("/{defect_id}/email-recipients")
async def get_email_recipients(
    defect_id: UUID,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    try:
        # Fetch defect with all relationships
        result = await db.execute(
            select(Defect)
            .where(Defect.id == defect_id)
            .options(
                selectinload(Defect.images),
                selectinload(Defect.threads).selectinload(Thread.attachments),
            )
        )
        defect = result.scalar_one_or_none()
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        vessel_result = await control_db.execute(
            select(Vessel).where(Vessel.imo == defect.vessel_imo)
        )
        vessel = vessel_result.scalars().first()
        vessel_name = vessel.name if vessel else defect.vessel_imo

        from app.models.associations import user_vessel_link

        stmt = (
            select(User)
            .join(user_vessel_link, User.id == user_vessel_link.c.user_id)
            .where(user_vessel_link.c.vessel_imo == defect.vessel_imo)
            .where(User.is_active == True)
        )
        users_result = await control_db.execute(stmt)
        users = users_result.scalars().all()
        recipients = [u.email for u in users if u.email]

        # Generate SAS URLs for before/after images
        before_images = []
        after_images = []
        for img in defect.images:
            url = generate_read_sas_url(img.blob_path)
            entry = {"file_name": img.file_name, "url": url}
            if img.image_type == "before":
                before_images.append(entry)
            else:
                after_images.append(entry)

        # Generate SAS URLs for thread attachments (skip system messages)
        thread_attachments = []
        for thread in defect.threads:
            if thread.is_system_message:
                continue
            for att in thread.attachments:
                url = generate_read_sas_url(att.blob_path)
                thread_attachments.append({"file_name": att.file_name, "url": url})

        # Build meaningful defect reference ID
        if defect.defect_number:
            defect_ref = defect.defect_number
        else:
            vessel_initials = "".join(w[0].upper() for w in vessel_name.split() if w)[
                :4
            ]
            date_part = (
                defect.date_identified.strftime("%Y%m%d")
                if defect.date_identified
                else defect.created_at.strftime("%Y%m%d")
            )
            short_uuid = str(defect.id)[:8]
            defect_ref = f"{vessel_initials}-{date_part}-{short_uuid}"

        subject, body_text = _build_email_content(defect, vessel_name, defect_ref)

        return {
            "recipients": recipients,
            "vessel_email": (
                vessel.vessel_email if vessel and vessel.vessel_email else None
            ),
            "defect_ref": defect_ref,
            "subject": subject,
            "body_text": body_text,
            "defect": {
                "id": str(defect.id),
                "title": defect.title,
                "vessel_name": vessel_name,
                "vessel_imo": defect.vessel_imo,
                "equipment_name": defect.equipment_name,
                "defect_source": (
                    defect.defect_source.value
                    if hasattr(defect.defect_source, "value")
                    else str(defect.defect_source)
                ),
                "priority": (
                    defect.priority.value
                    if hasattr(defect.priority, "value")
                    else str(defect.priority)
                ),
                "status": (
                    defect.status.value
                    if hasattr(defect.status, "value")
                    else str(defect.status)
                ),
                "responsibility": defect.responsibility or "N/A",
                "description": defect.description,
                "date_identified": (
                    defect.date_identified.strftime("%Y-%m-%d")
                    if defect.date_identified
                    else "N/A"
                ),
                "target_close_date": (
                    defect.target_close_date.strftime("%Y-%m-%d")
                    if defect.target_close_date
                    else "N/A"
                ),
            },
            "attachments": {
                "before_images": before_images,
                "after_images": after_images,
                "thread_attachments": thread_attachments,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[email-recipients] Failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")


@router.post("/{defect_id}/draft-email")
async def create_email_draft(
    defect_id: UUID,
    request: EmailDraftRequest = EmailDraftRequest(),
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Creates an Outlook draft email with defect details and
    actual file attachments (before/after images + thread attachments).
    Accepts optional overrides for to_emails, cc_emails, subject, body_text
    (used when the vessel user edits the email in-app before saving).
    """
    try:
        # 1. Fetch defect with all relationships
        result = await db.execute(
            select(Defect)
            .where(Defect.id == defect_id)
            .options(
                selectinload(Defect.images),
                selectinload(Defect.threads).selectinload(Thread.attachments),
            )
        )
        defect = result.scalar_one_or_none()
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        # 2. Get vessel name
        vessel_result = await control_db.execute(
            select(Vessel).where(Vessel.imo == defect.vessel_imo)
        )
        vessel = vessel_result.scalars().first()
        vessel_name = vessel.name if vessel else defect.vessel_imo
        vessel_email = vessel.vessel_email if vessel and vessel.vessel_email else None

        # 3. Get recipients
        from app.models.associations import user_vessel_link

        stmt = (
            select(User)
            .join(user_vessel_link, User.id == user_vessel_link.c.user_id)
            .where(user_vessel_link.c.vessel_imo == defect.vessel_imo)
            .where(User.is_active == True)
        )
        users_result = await control_db.execute(stmt)
        users = users_result.scalars().all()
        recipients = [u.email for u in users if u.email]

        # Use stored defect_number, fallback to generating if null (old records)
        if defect.defect_number:
            defect_ref = defect.defect_number
        else:
            vessel_initials = "".join(w[0].upper() for w in vessel_name.split() if w)[
                :4
            ]
            date_part = (
                defect.date_identified.strftime("%Y%m%d")
                if defect.date_identified
                else defect.created_at.strftime("%Y%m%d")
            )
            short_uuid = str(defect.id)[:8]
            defect_ref = f"{vessel_initials}-{date_part}-{short_uuid}"

        # 5. Build email subject/body (use overrides if provided by vessel in-app editor)
        auto_subject, auto_body = _build_email_content(defect, vessel_name, defect_ref)
        subject = request.subject if request.subject is not None else auto_subject
        body_text = request.body_text if request.body_text is not None else auto_body
        recipients = request.to_emails if request.to_emails is not None else recipients
        cc_list = request.cc_emails if request.cc_emails is not None else ([vessel_email] if vessel_email else [])

        # 5. Download blobs and build attachments list
        attachments = []
        MAX_ATTACH_SIZE = 3 * 1024 * 1024  # 3MB per file limit
        total_size = 0
        MAX_TOTAL_SIZE = 20 * 1024 * 1024  # 20MB total limit

        # Before/After images
        for img in defect.images:
            try:
                data = download_blob_bytes(img.blob_path)
                if total_size + len(data) > MAX_TOTAL_SIZE:
                    logger.warning(
                        f"Skipping {img.file_name} — total size limit reached"
                    )
                    continue
                if len(data) > MAX_ATTACH_SIZE:
                    logger.warning(f"Skipping {img.file_name} — file too large")
                    continue
                attachments.append(
                    {
                        "name": f"[{img.image_type.upper()}] {img.file_name}",
                        "content_type": "image/jpeg",
                        "data": data,
                    }
                )
                total_size += len(data)
            except Exception as e:
                logger.warning(f"Could not download image {img.file_name}: {e}")

        # Thread attachments
        for thread in defect.threads:
            if thread.is_system_message:
                continue
            for att in thread.attachments:
                try:
                    data = download_blob_bytes(att.blob_path)
                    if total_size + len(data) > MAX_TOTAL_SIZE:
                        logger.warning(
                            f"Skipping {att.file_name} — total size limit reached"
                        )
                        continue
                    if len(data) > MAX_ATTACH_SIZE:
                        logger.warning(f"Skipping {att.file_name} — file too large")
                        continue
                    attachments.append(
                        {
                            "name": att.file_name,
                            "content_type": att.content_type
                            or "application/octet-stream",
                            "data": data,
                        }
                    )
                    total_size += len(data)
                except Exception as e:
                    logger.warning(
                        f"Could not download attachment {att.file_name}: {e}"
                    )

        # 6. Create Outlook draft via Graph API
        # subject, body_text, recipients, cc_list are set in step 5 (with user overrides applied)
        web_link = await create_outlook_draft(
            user_email=current_user.email,
            # user_email="techdevops@ozellar.com",  # was hardcoded for testing
            to_emails=recipients,
            cc_emails=cc_list,
            subject=subject,
            body_text=body_text,
            attachments=attachments,
        )

        return {
            "success": True,
            "web_link": web_link,
            "attachment_count": len(attachments),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[draft-email] Failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create draft: {str(e)}")


def get_image_data(blob_path):
    """
    Downloads blob, resizes to fit strictly within 180x180px thumbnail.
    Returns (BytesIO, width, height) for perfect centering.
    """
    if not blob_path:
        return None, 0, 0
    try:
        blob_client = get_container_client().get_blob_client(blob=blob_path)
        if not blob_client.exists():
            return None, 0, 0

        stream = io.BytesIO()
        blob_data = blob_client.download_blob()
        blob_data.readinto(stream)
        stream.seek(0)

        img = PILImage.open(stream)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        # ✅ FORCE STANDARD SIZE: 180x180 max
        # This keeps file size low and layout consistent
        img.thumbnail((1000, 1000), PILImage.Resampling.LANCZOS)

        # Get actual new dimensions
        width, height = img.size

        output_buffer = io.BytesIO()
        img.save(output_buffer, format="JPEG", quality=95, subsampling=0)
        output_buffer.seek(0)

        return output_buffer, width, height

    except Exception as e:
        logger.error(f"Image Error: {e}")
        return None, 0, 0


async def get_vessel_name(control_db: AsyncSession, vessel_imo: str) -> str:
    """Fetch vessel name from workplace_control by IMO."""
    result = await control_db.execute(select(Vessel).where(Vessel.imo == vessel_imo))
    vessel = result.scalars().first()
    return vessel.name if vessel else vessel_imo


# =============================================================================
# GET ALL DEFECTS
# =============================================================================
@router.get("/", response_model=list[DefectResponse])
async def get_defects(
    vessel_imo: str | None = None,
    status: str | None = None,
    priority: str | None = None,
    defect_source: str | None = None,
    equipment_name: str | None = None,
    is_owner: bool | None = None,
    is_flagged: bool | None = None,  # ✅ Add this
    is_dd: bool | None = None,  # ✅ Add this
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    query = (
        select(Defect)
        .options(selectinload(Defect.pr_entries))  # ← vessel removed
        .where(Defect.is_deleted == False)
        .order_by(Defect.date_identified.desc(), Defect.created_at.desc())
    )

    # Vessel users only see defects for their own vessel(s)
    if current_user.role == UserRole.VESSEL:
        user_vessel_imos = [v.imo for v in current_user.vessels]
        if not user_vessel_imos:
            raise HTTPException(status_code=403, detail="No vessels assigned")
        query = query.where(Defect.vessel_imo.in_(user_vessel_imos))

    elif current_user.role == UserRole.SHORE:
        user_vessel_imos = [v.imo for v in current_user.vessels]
        if not user_vessel_imos:
            raise HTTPException(status_code=403, detail="No vessels assigned")
        query = query.where(Defect.vessel_imo.in_(user_vessel_imos))
    elif vessel_imo:
        query = query.where(Defect.vessel_imo == vessel_imo)

    if status:
        try:
            query = query.where(Defect.status == DefectStatus(status))
        except ValueError:
            pass
    if priority:
        try:
            query = query.where(Defect.priority == DefectPriority(priority))
        except ValueError:
            pass
    if defect_source:
        query = query.where(Defect.defect_source == defect_source)
    if equipment_name:
        query = query.where(Defect.equipment_name.ilike(f"%{equipment_name}%"))
    if is_owner is not None:
        query = query.where(Defect.is_owner == is_owner)
    if is_dd is not None:
        query = query.where(Defect.is_dd == is_dd)

    result = await db.execute(query)
    defects = result.scalars().all()

    # Bulk fetch vessel names from control DB (one query, not N queries)
    vessel_imos = list(set(d.vessel_imo for d in defects))
    vessel_result = await control_db.execute(
        select(Vessel).where(Vessel.imo.in_(vessel_imos))
    )
    vessel_map = {v.imo: v.name for v in vessel_result.scalars().all()}
    flag_result = await db.execute(
        select(UserDefectFlag.defect_id).where(
            UserDefectFlag.user_id == current_user.id,
            UserDefectFlag.defect_id.in_([d.id for d in defects])
        )
    )
    flagged_ids = {row[0] for row in flag_result.all()}

    for defect in defects:
        # Use __dict__ to avoid mutating the ORM collection (prevents SQLAlchemy
        # from nullifying defect_id on removed items at session flush)
        defect.__dict__['pr_entries'] = [pr for pr in defect.pr_entries if not pr.is_deleted]
        defect.vessel_name = vessel_map.get(defect.vessel_imo, defect.vessel_imo)
        defect.__dict__['is_flagged'] = defect.id in flagged_ids

    return defects


# =============================================================================
# EXCEL EXPORT — Multi-select filters, IST dates, embedded images
# =============================================================================
@router.get("/export", response_class=StreamingResponse)
async def export_defects(
    #  MULTI-SELECT FILTERS (Accept Lists)
    vessel_imo: list[str] | None = Query(None),
    status: list[str] | None = Query(None),
    priority: list[str] | None = Query(None),
    defect_source: list[str] | None = Query(None),
    equipment_name: list[str] | None = Query(None),
    #  TEXT SEARCH FILTERS
    description: str | None = None,
    pr_number: str | None = None,
    #  DATE FILTERS
    date_identified_from: str | None = None,
    date_identified_to: str | None = None,
    target_close_date: str | None = None,
    # OTHER FILTERS
    is_owner: bool | None = None,
    is_dd: list[str] | None = Query(None),
    # COLUMN CUSTOMIZATION
    visible_columns: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    control_db: AsyncSession = Depends(get_control_db),
):
    """
    ✅ PRODUCTION-READY Excel Export with ALL fixes:
    - Multi-select filtering using IN operator (not ILIKE)
    - Date formatting fix (prevents -1 day offset)
    - Sheet 2 cleaned up (removed Priority, Status, Closure Remarks)
    - Standardized image sizing (180x180, centered in 200x200 cells)
    """
    try:
        logger.info(f"📊 Exporting defects with filters...")
        logger.info(f"   Vessels: {vessel_imo}")
        logger.info(f"   Equipment: {equipment_name}")
        logger.info(f"   Status: {status}")
        logger.info(f"   Priority: {priority}")
        logger.info(f"   Source: {defect_source}")
        logger.info(f"   Description search: {description}")
        logger.info(f"   PR search: {pr_number}")
        logger.info(f"   Date range: {date_identified_from} to {date_identified_to}")
        logger.info(f"   Deadline before: {target_close_date}")

        # ===== 1. BUILD QUERY WITH FILTERS =====
        query = (
            select(Defect)
            .options(
                selectinload(Defect.pr_entries),
                selectinload(Defect.images),  # ✅ Load images for Sheet 2
            )
            .where(Defect.is_deleted == False)
            .order_by(Defect.date_identified.desc())
        )

        # ✅ NEW: Apply role-based vessel filtering
        if current_user.role == UserRole.VESSEL:
            user_vessel_imos = [v.imo for v in current_user.vessels]
            if not user_vessel_imos:
                raise HTTPException(status_code=403, detail="No vessels assigned")
            query = query.where(Defect.vessel_imo.in_(user_vessel_imos))

        elif current_user.role == UserRole.SHORE:
            user_vessel_imos = [v.imo for v in current_user.vessels]
            if not user_vessel_imos:
                raise HTTPException(status_code=403, detail="No vessels assigned")
            query = query.where(Defect.vessel_imo.in_(user_vessel_imos))

        # ADMIN users see all defects (no filtering)

        # ✅ FIXED: Multi-select filters using IN operator (exact match)
        if vessel_imo and len(vessel_imo) > 0:
            query = query.where(Defect.vessel_imo.in_(vessel_imo))

        # ✅ FIXED: Equipment filter using IN operator (exact match)
        if equipment_name and len(equipment_name) > 0:
            query = query.where(Defect.equipment_name.in_(equipment_name))

        if status and len(status) > 0:
            try:
                status_enums = [DefectStatus(s.upper()) for s in status]
                query = query.where(Defect.status.in_(status_enums))
            except ValueError:
                pass

        if priority and len(priority) > 0:
            try:
                priority_enums = [DefectPriority(p.upper()) for p in priority]
                query = query.where(Defect.priority.in_(priority_enums))
            except ValueError:
                pass

        if defect_source and len(defect_source) > 0:
            query = query.where(Defect.defect_source.in_(defect_source))

        # Text search filters
        if description:
            query = query.where(Defect.description.ilike(f"%{description}%"))

        # Date filters
        if date_identified_from:
            try:
                from_date = datetime.strptime(date_identified_from, "%Y-%m-%d")
                query = query.where(Defect.date_identified >= from_date)
            except ValueError:
                pass

        if date_identified_to:
            try:
                # End of day for inclusive filtering
                to_date = datetime.strptime(date_identified_to, "%Y-%m-%d").replace(
                    hour=23, minute=59, second=59
                )
                query = query.where(Defect.date_identified <= to_date)
            except ValueError:
                pass

        if target_close_date:
            try:
                deadline = datetime.strptime(target_close_date, "%Y-%m-%d")
                query = query.where(Defect.target_close_date <= deadline)
            except ValueError:
                pass

        if is_owner is not None:
            query = query.where(Defect.is_owner == is_owner)


        if is_dd and len(is_dd) > 0:
            if "true" in is_dd and "false" not in is_dd:
                query = query.where(Defect.is_dd == True)
            elif "false" in is_dd and "true" not in is_dd:
                query = query.where(Defect.is_dd == False)

        # Execute query
        result = await db.execute(query)
        defects = result.scalars().all()
        vessel_imos = list(set(d.vessel_imo for d in defects))
        if vessel_imos:
            vessel_result = await control_db.execute(
                select(Vessel).where(Vessel.imo.in_(vessel_imos))
            )
            vessel_map = {v.imo: v.name for v in vessel_result.scalars().all()}
        else:
            vessel_map = {}

        # Filter by PR Number (post-query filter for complex join)
        if pr_number:
            defects = [
                d
                for d in defects
                if any(
                    pr_number.lower() in pr.pr_number.lower()
                    for pr in d.pr_entries
                    if not pr.is_deleted
                )
            ]
        
        logger.info(f"✅ Found {len(defects)} defects matching filters")

        # ===== 2. COLUMN SELECTION =====
        if visible_columns:
            column_keys = [
                c.strip() for c in visible_columns.split(",") if c.strip() in COLUMN_MAP
            ]
            if "closure_remarks" not in column_keys:
                column_keys.append("closure_remarks")
        else:
            # Default columns if none specified
            column_keys = [
                "date",
                "deadline",
                "source",
                "equipment",
                "description",
                "priority",
                "status",
                "owner",
                "pr_details",
                "closure_remarks",
            ]

        logger.info(f"📋 Exporting columns: {column_keys}")

        # ===== 3. CREATE EXCEL WORKBOOK =====
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {"in_memory": True})

        # Define formats
        header_fmt = wb.add_format(
            {
                "bold": True,
                "bg_color": "#366092",
                "font_color": "white",
                "align": "center",
                "valign": "vcenter",
                "border": 1,
            }
        )

        text_fmt = wb.add_format({"text_wrap": True, "valign": "top", "border": 1})

        center_fmt = wb.add_format({"align": "center", "valign": "top", "border": 1})

        # ================= SHEET 1: DEFECT LIST =================
        ws1 = wb.add_worksheet("Defect List")

        # IMPORTANT: Unlock ALL cells by default first — then re-lock only col 0
        # xlsxwriter locks everything by default; we must explicitly unlock the rest
        ws1.set_column(0, 0, 18)   # Defect ID column width

        locked_fmt = wb.add_format({
            "text_wrap": True, "valign": "top", "border": 1,
            "bg_color": "#e2e8f0", "font_color": "#475569",
            "bold": False, "locked": True,
            "italic": True,
        })
        text_fmt = wb.add_format({
            "text_wrap": True, "valign": "top", "border": 1, "locked": False
        })
        center_fmt = wb.add_format({
            "align": "center", "valign": "top", "border": 1, "locked": False
        })

        # col 1 gets its width here too, since this call overrides the earlier set_column
        ws1.set_column(1, 1, 20, wb.add_format({"locked": False}))   # Vessel — unlocked, width 20
        ws1.set_column(2, 50, None, wb.add_format({"locked": False})) # All other data cols — unlocked

        ws1.write(0, 0, "Defect ID", header_fmt)
        ws1.write(0, 1, "Vessel Name", header_fmt)

        # Write dynamic column headers
        unlocked_col_fmt = wb.add_format({"locked": False})
        col = 2
        for key in column_keys:
            ws1.write(0, col, COLUMN_MAP[key][0], header_fmt)
            width = 50 if key == "description" else 20
            ws1.set_column(col, col, width, unlocked_col_fmt)
            col += 1

        # Write data rows
        for i, defect in enumerate(defects):
            row = i + 1
            vessel_name = defect.vessel_imo

            defect_id_val = defect.defect_number if defect.defect_number else f"#{i + 1}"
            ws1.write(row, 0, defect_id_val, locked_fmt)
            ws1.write(
                row, 1, vessel_map.get(defect.vessel_imo, defect.vessel_imo), text_fmt
            )

            col = 2
            for key in column_keys:
                # ✅ FIX: Convert dates to simple strings to prevent timezone shifts
                if key == "date":
                    val = (
                        defect.date_identified.astimezone(IST).strftime("%Y-%m-%d")
                        if defect.date_identified
                        else "-"
                    )
                    ws1.write(row, col, val, center_fmt)

                elif key == "deadline":
                    val = (
                        defect.target_close_date.astimezone(IST).strftime("%Y-%m-%d")
                        if defect.target_close_date
                        else "-"
                    )
                    ws1.write(row, col, val, center_fmt)

                elif key == "closure_remarks":
                    val = COLUMN_MAP[key][1](defect, i + 1)
                    ws1.write(row, col, val, text_fmt)  # Use text_fmt for wrapping
                else:
                    # Use standard mapper for other columns
                    val = COLUMN_MAP[key][1](defect, i + 1)
                    fmt = (
                        text_fmt if key in ["description", "equipment"] else center_fmt
                    )
                    ws1.write(row, col, val, fmt)
                col += 1

        ws1.freeze_panes(1, 0)
        ws1.autofilter(0, 0, len(defects), col - 1)

        # ===== ADD IMPORT-COMPATIBLE DROPDOWNS =====
        # Build a hidden Lists sheet with dropdown sources
        ws_lists = wb.add_worksheet("_Lists")
        ws_lists.hide()

        # from app.models.enums import DefectSource, DefectPriority, DefectStatus

        source_options = [e.value for e in DefectSource]
        priority_options = [e.name for e in DefectPriority]
        status_options = [e.name for e in DefectStatus]

        for i, opt in enumerate(EQUIPMENT_OPTIONS):
            ws_lists.write(i, 0, opt)
        for i, opt in enumerate(source_options):
            ws_lists.write(i, 1, opt)
        for i, opt in enumerate(priority_options):
            ws_lists.write(i, 2, opt)
        for i, opt in enumerate(status_options):
            ws_lists.write(i, 3, opt)

        # Map column keys to their Excel column index (col 0=DefectID, 1=Vessel, 2+ dynamic)
        key_to_col = {key: 2 + idx for idx, key in enumerate(column_keys)}

        if "source" in key_to_col:
            c = key_to_col["source"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "list",
                "source": f"=_Lists!$B$1:$B${len(source_options)}",
            })

        if "equipment" in key_to_col:
            c = key_to_col["equipment"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "list",
                "source": f"=_Lists!$A$1:$A${len(EQUIPMENT_OPTIONS)}",
            })

        if "priority" in key_to_col:
            c = key_to_col["priority"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "list",
                "source": f"=_Lists!$C$1:$C${len(priority_options)}",
            })

        if "status" in key_to_col:
            c = key_to_col["status"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "list",
                "source": f"=_Lists!$D$1:$D${len(status_options)}",
            })

        for flag_key in ["dd"]:
            if flag_key in key_to_col:
                c = key_to_col[flag_key]
                ws1.data_validation(1, c, len(defects) + 100, c, {
                    "validate": "list",
                    "source": ["Yes", "No"],
                })

        if "owner" in key_to_col:
            c = key_to_col["owner"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "list",
                "source": ["Owner", "Not Owner"],
            })

        if "date" in key_to_col:
            c = key_to_col["date"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "date",
                "criteria": "between",
                "minimum": datetime(2000, 1, 1).date(),
                "maximum": datetime(2100, 12, 31).date(),
                "input_message": "Format: YYYY-MM-DD",
            })

        if "deadline" in key_to_col:
            c = key_to_col["deadline"]
            ws1.data_validation(1, c, len(defects) + 100, c, {
                "validate": "date",
                "criteria": "between",
                "minimum": datetime(2000, 1, 1).date(),
                "maximum": datetime(2100, 12, 31).date(),
                "input_message": "Format: YYYY-MM-DD",
            })
        # ===== END DROPDOWNS =====

        # Protect sheet but allow editing all cells except Defect ID (col 0)
        ws1.protect('', {
            'sheet':                  True,
            'select_locked_cells':    True,
            'select_unlocked_cells':  True,
            'format_cells':           False,
            'format_columns':         False,
            'format_rows':            False,
            'insert_columns':         False,
            'insert_rows':            False,
            'delete_columns':         False,
            'delete_rows':            False,
            'sort':                   True,
            'autofilter':             True,
            'objects':                False,
            'scenarios':              False,
        })
        # ================= SHEET 2: DETAILED REPORT (IMAGES) =================
        ws2 = wb.add_worksheet("Detailed Report")

        # ✅ CLEANED UP: Removed Priority, Status, and Closure Remarks
        headers2 = [
            "S.No",
            "Vessel",
            "Report Date",
            "Area of Concern",
            "Defect Description",
            "Before Image 1",
            "Before Image 2",
            "After Image 1",
            "After Image 2",
        ]

        # Set column widths
        ws2.set_column("A:A", 8)  # S.No
        ws2.set_column("B:B", 20)  # Vessel
        ws2.set_column("C:C", 15)  # Report Date
        ws2.set_column("D:D", 20)  # Area of Concern
        ws2.set_column("E:E", 40)  # Description
        ws2.set_column("F:I", 50)  # ✅ Images (Now F to I)

        # Write headers
        for c, h in enumerate(headers2):
            ws2.write(0, c, h, header_fmt)

        # Write defect rows with images
        for i, defect in enumerate(defects):
            row = i + 1

            # ✅ FORCE ROW HEIGHT: 150 points (~200px)
            ws2.set_row(row, 255)
            report_date = (
                defect.date_identified.astimezone(IST).strftime("%Y-%m-%d")
                if defect.date_identified
                else "-"
            )
            # Write text columns
            ws2.write(row, 0, i + 1, center_fmt)  # Col 0: S.No
            ws2.write(
                row, 1, vessel_map.get(defect.vessel_imo, defect.vessel_imo), text_fmt
            )  # Col 1: Vessel     # Col 1: Vessel
            ws2.write(row, 2, report_date, center_fmt)  # Col 2: Report Date
            ws2.write(row, 3, defect.equipment_name, text_fmt)
            ws2.write(row, 4, defect.description, text_fmt)

            # Sort images by creation date
            before = sorted(
                [img for img in defect.images if img.image_type == "before"],
                key=lambda x: x.created_at,
                reverse=True,
            )
            after = sorted(
                [img for img in defect.images if img.image_type == "after"],
                key=lambda x: x.created_at,
                reverse=True,
            )

            def place_image(col_idx, img_obj):
                """Insert image centered in cell with guaranteed no overlap"""
                if not img_obj:
                    ws2.write(row, col_idx, "No Image", center_fmt)
                    return

                buf, w, h = get_image_data(img_obj.blob_path)
                if buf:
                    # ✅ THE FIX:
                    # Container is 340px high.
                    # We force the image to fit inside a 310px box.
                    container_h = 340
                    container_w = 350  # Based on Column Width 50
                    max_target = 310

                    # Calculate scale so the largest side is 310px
                    scale = max_target / max(w, h)

                    # Calculate offsets to center the image within the container
                    x_off = (container_w - (w * scale)) / 2
                    y_off = (container_h - (h * scale)) / 2

                    try:
                        ws2.insert_image(
                            row,
                            col_idx,
                            img_obj.file_name,
                            {
                                "image_data": buf,
                                "x_scale": scale,
                                "y_scale": scale,
                                "x_offset": x_off,
                                "y_offset": y_off,
                                "object_position": 1,  # Move and size with cells
                            },
                        )
                    except Exception as img_err:
                        logger.error(f"Failed to insert image: {img_err}")
                        ws2.write(row, col_idx, "Err", center_fmt)  # Insert images

            place_image(5, before[0] if len(before) > 0 else None)
            place_image(6, before[1] if len(before) > 1 else None)
            place_image(7, after[0] if len(after) > 0 else None)
            place_image(8, after[1] if len(after) > 1 else None)

        # ===== 5. FINALIZE AND RETURN =====
        wb.close()
        output.seek(0)

        filename = f"Defect_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

        logger.info(f"✅ Excel export completed: {filename}")

        return StreamingResponse(
            output,
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        logger.error(f"❌ Export failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


# =============================================================================
# SAS URL GENERATION
# =============================================================================
@router.get("/sas")
async def get_upload_sas(blobName: str, current_user: User = Depends(get_current_user)):
    """Generate upload SAS URL for direct browser-to-blob upload."""
    try:
        logger.info(f"📤 Generating upload SAS for: {blobName}")
        url = generate_write_sas_url(blobName)
        return {"url": url}
    except Exception as e:
        logger.error(f"❌ Error generating upload SAS: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Failed to generate upload URL: {str(e)}"
        )


# =============================================================================
# VALIDATE IMAGES BEFORE CLOSE
# =============================================================================
@router.get("/{defect_id}/validate-images")
async def validate_defect_images(
    defect_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Validates that all mandatory images are present before allowing closure.
    Called by the UI before showing the closure form.
    """
    try:
        defect = await db.get(Defect, defect_id)
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        missing_images = []

        if defect.before_image_required:
            result = await db.execute(
                select(DefectImage).where(
                    DefectImage.defect_id == defect_id,
                    DefectImage.image_type == "before",
                )
            )
            if not result.scalars().all():
                missing_images.append("⚠️ Before images are mandatory but none uploaded")

        if defect.after_image_required:
            result = await db.execute(
                select(DefectImage).where(
                    DefectImage.defect_id == defect_id,
                    DefectImage.image_type == "after",
                )
            )
            if not result.scalars().all():
                missing_images.append("⚠️ After images are mandatory but none uploaded")

        return {"can_close": len(missing_images) == 0, "missing_images": missing_images}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Validation error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# IMPORT TEMPLATE — SHORE (Vessel Name dropdown populated live from DB)
# =============================================================================
@router.get("/import-template")
async def download_import_template(
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
):
    """
    Shore-side import template.
    Includes Vessel Name dropdown populated live from the database.
    Uses "Area of Concern" as the equipment column header.
    """
    vessel_stmt = await control_db.execute(select(Vessel.name).order_by(Vessel.name))
    vessel_names = [v[0] for v in vessel_stmt.all()]

    source_options = [e.value for e in DefectSource]
    priority_options = [e.name for e in DefectPriority]
    status_options = [e.name for e in DefectStatus]

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("Defects")
    data_ws = workbook.add_worksheet("Lists")
    data_ws.hide()

    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#366092",
            "font_color": "white",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )

    # Populate hidden Lists sheet for dropdown sources
    for i, name in enumerate(vessel_names):
        data_ws.write(i, 0, name)
    for i, opt in enumerate(EQUIPMENT_OPTIONS):
        data_ws.write(i, 1, opt)
    for i, opt in enumerate(source_options):
        data_ws.write(i, 2, opt)
    for i, opt in enumerate(priority_options):
        data_ws.write(i, 3, opt)
    for i, opt in enumerate(status_options):
        data_ws.write(i, 4, opt)

    headers = [
        "Vessel Name",
        "Report Date",
        "Deadline",
        "Defect Source",
        "Area of Concern",
        "Description",
        "Priority",
        "Status",
        "PR Number",
        "Dry Dock",
    ]
    for col, text in enumerate(headers):
        ws.write(0, col, text, header_fmt)
        ws.set_column(col, col, 22)

    # Dropdowns
    ws.data_validation(
        1,
        0,
        1000,
        0,
        {
            "validate": "list",
            "source": f"=Lists!$A$1:$A${len(vessel_names)}",
            "ignore_blank": True,
        },
    )
    ws.data_validation(
        1,
        3,
        1000,
        3,
        {"validate": "list", "source": f"=Lists!$C$1:$C${len(source_options)}"},
    )
    ws.data_validation(
        1,
        4,
        1000,
        4,
        {"validate": "list", "source": f"=Lists!$B$1:$B${len(EQUIPMENT_OPTIONS)}"},
    )
    ws.data_validation(
        1,
        6,
        1000,
        6,
        {"validate": "list", "source": f"=Lists!$D$1:$D${len(priority_options)}"},
    )
    ws.data_validation(
        1,
        7,
        1000,
        7,
        {"validate": "list", "source": f"=Lists!$E$1:$E${len(status_options)}"},
    )

    # Date columns with validation
    date_hint_fmt = workbook.add_format({"num_format": "yyyy-mm-dd"})
    for col_idx in [1, 2]:
        ws.set_column(col_idx, col_idx, 20, date_hint_fmt)
        ws.data_validation(
            1,
            col_idx,
            1000,
            col_idx,
            {
                "validate": "date",
                "criteria": "between",
                "minimum": datetime(2000, 1, 1).date(),
                "maximum": datetime(2100, 12, 31).date(),
                "input_title": "Date Required",
                "input_message": "Please enter date as YYYY-MM-DD",
                "error_title": "Invalid Date",
                "error_message": "Please use the format: 2023-12-31",
            },
        )
    start_col = 8 if "Vessel Name" in headers else 7
    for col_idx in [start_col, start_col + 1]:
        ws.data_validation(
            1,
            col_idx,
            1000,
            col_idx,
            {
                "validate": "list",
                "source": ["Yes", "No"],
                "input_title": "Select Option",
                "input_message": "Please select Yes or No",
            },
        )
        ws.data_validation(1, 9, 1000, 9, {
        "validate": "list",
        "source": ["Yes", "No"],
        "input_title": "Select Option",
        "input_message": "Please select Yes or No",
        })

    workbook.close()
    output.seek(0)
    filename = f"Defect_Import_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =============================================================================
# IMPORT TEMPLATE — VESSEL (No Vessel Name column — auto-filled on import)
# =============================================================================
@router.get("/import-template-vessel")
async def download_import_template_vessel(
    current_user: User = Depends(get_current_user),
):
    """
    Vessel-side import template.
    EXCLUDES Vessel Name column — auto-filled from logged-in user's vessel during import.
    Uses "Area of Concern" as the equipment column header.
    """
    source_options = [e.value for e in DefectSource]
    priority_options = [e.name for e in DefectPriority]
    status_options = [e.name for e in DefectStatus]

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    ws = workbook.add_worksheet("Defects")
    data_ws = workbook.add_worksheet("Lists")
    data_ws.hide()

    # Hidden Lists sheet (no vessel column — vessel is auto-assigned)
    for i, opt in enumerate(EQUIPMENT_OPTIONS):
        data_ws.write(i, 0, opt)
    for i, opt in enumerate(source_options):
        data_ws.write(i, 1, opt)
    for i, opt in enumerate(priority_options):
        data_ws.write(i, 2, opt)
    for i, opt in enumerate(status_options):
        data_ws.write(i, 3, opt)

    header_fmt = workbook.add_format(
        {
            "bold": True,
            "bg_color": "#366092",
            "font_color": "white",
            "border": 1,
            "align": "center",
            "valign": "vcenter",
        }
    )

    headers = [
        "Report Date",
        "Deadline",
        "Defect Source",
        "Area of Concern",
        "Description",
        "Priority",
        "Status",
        "PR Number",
        "Dry Dock",  # ✅ Added
    ]
    for col, text in enumerate(headers):
        ws.write(0, col, text, header_fmt)
        ws.set_column(col, col, 22)

    ws.data_validation(
        1,
        2,
        1000,
        2,
        {"validate": "list", "source": f"=Lists!$B$1:$B${len(source_options)}"},
    )
    ws.data_validation(
        1,
        3,
        1000,
        3,
        {"validate": "list", "source": f"=Lists!$A$1:$A${len(EQUIPMENT_OPTIONS)}"},
    )
    ws.data_validation(
        1,
        5,
        1000,
        5,
        {"validate": "list", "source": f"=Lists!$C$1:$C${len(priority_options)}"},
    )
    ws.data_validation(
        1,
        6,
        1000,
        6,
        {"validate": "list", "source": f"=Lists!$D$1:$D${len(status_options)}"},
    )

    date_fmt = workbook.add_format({"num_format": "yyyy-mm-dd"})
    ws.set_column(0, 1, 20, date_fmt)
    for col_idx in [0, 1]:
        ws.data_validation(
            1,
            col_idx,
            1000,
            col_idx,
            {
                "validate": "date",
                "criteria": "between",
                "minimum": datetime(2000, 1, 1).date(),
                "maximum": datetime(2100, 12, 31).date(),
                "input_message": "Format: YYYY-MM-DD",
            },
        )
    for col_idx in [8]:
        ws.data_validation(
            1,
            col_idx,
            1000,
            col_idx,
            {
                "validate": "list",
                "source": ["Yes", "No"],
                "input_title": "Select Option",
                "input_message": "Please select Yes or No",
            },
        )

    workbook.close()
    output.seek(0)
    filename = f"Defect_Import_Template_Vessel_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =============================================================================
# BULK IMPORT
# Uses openpyxl + "Area of Concern" column name.
# Batch insert (1000-row chunks) for speed — much faster than ORM loop.
# SyncQueue entries written in the same batch on vessel/offline mode.
# Comma-separated PR numbers split into individual PrEntry rows.
# Full duplicate check: vessel + equipment + description + date + source.
# =============================================================================
@router.post("/import")
async def import_defects(
    file: UploadFile = File(...),
    skip_errors: bool = Query(True),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Bulk import/update defects from Excel.
 
    Rows WITH a valid Defect ID (matching defect_number in DB) → UPDATE existing defect.
    Rows WITHOUT a Defect ID (or with a placeholder like #1) → CREATE new defect.
 
    Returns separate counts for created vs updated rows.
    """
    try:
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
 
        contents = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        sheet = workbook.active
 
        headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
 
        required_columns = [
            "Area of Concern",
            "Description",
            "Defect Source",
            "Priority",
            "Status",
            "Report Date",
            "Deadline",
        ]
        if current_user.role != UserRole.VESSEL:
            required_columns.insert(0, "Vessel Name")
 
        for col in required_columns:
            if col not in headers:
                raise HTTPException(
                    status_code=400, detail=f"Missing required column: '{col}'"
                )
 
        def get_enum_value(enum_class, value):
            if not value:
                return None
            val_str = str(value).strip().lower()
            for e in enum_class:
                if e.name.lower() == val_str or e.value.lower() == val_str:
                    return e
            return None
 
        def parse_date(val):
            if isinstance(val, datetime):
                return val
            if not val:
                return None
            try:
                return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d")
            except Exception:
                return None
 
        # Track objects to persist — kept separate so we can report counts accurately
        defects_to_insert = []    # new Defect ORM objects
        defects_to_update = []    # existing Defect ORM objects (already fetched + mutated)
        prs_to_insert = []        # new PrEntry ORM objects (for both create and update paths)
        syncs_to_insert = []      # SyncQueue entries
 
        created_count = 0
        updated_count = 0
        errors = []
 
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            try:
                # Skip completely blank rows (Excel trailing empty rows)
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
 
                row_data = dict(zip(headers, row))
 
                # Treat "none" / "None" string as missing
                missing_fields = [
                    f
                    for f in required_columns
                    if row_data.get(f) is None
                    or str(row_data.get(f)).strip() == ""
                    or str(row_data.get(f)).strip().lower() == "none"
                ]
                if missing_fields:
                    raise ValueError(
                        f"Missing required field(s): {', '.join(missing_fields)}"
                    )
 
                # ── Resolve vessel ─────────────────────────────────────────
                if current_user.role == UserRole.VESSEL:
                    vessel = current_user.vessels[0]
                else:
                    v_name = (
                        str(row_data["Vessel Name"])
                        .encode("ascii", "ignore")
                        .decode()
                        .strip()
                    )
                    v_res = await control_db.execute(
                        select(Vessel).where(Vessel.name.ilike(v_name))
                    )
                    vessel = v_res.scalars().first()
                    if not vessel:
                        raise ValueError(f"Vessel '{v_name}' not found in database")
 
                # ── Parse enums and dates ──────────────────────────────────
                source_enum = get_enum_value(DefectSource, row_data["Defect Source"])
                priority_enum = get_enum_value(DefectPriority, row_data["Priority"])
                status_enum = get_enum_value(DefectStatus, row_data["Status"])
                date_id = parse_date(row_data["Report Date"])
                date_dl = parse_date(row_data["Deadline"])
 
                validation_errors = []
                if not source_enum:
                    validation_errors.append(
                        f"Invalid Source: {row_data.get('Defect Source')}"
                    )
                if not priority_enum:
                    validation_errors.append(
                        f"Invalid Priority: {row_data.get('Priority')}"
                    )
                if not status_enum:
                    validation_errors.append(
                        f"Invalid Status: {row_data.get('Status')}"
                    )
                if not date_id:
                    validation_errors.append("Invalid Report Date format")
                if not date_dl:
                    validation_errors.append("Invalid Deadline format")
                if validation_errors:
                    raise ValueError(" | ".join(validation_errors))
 
                equip_str = str(row_data["Area of Concern"]).strip()
                desc_str = str(row_data["Description"]).strip()
 
                # FIX: strip whitespace before boolean comparisons
                is_owner_val = (
                    str(row_data.get("Owner", "")).strip().lower() == "owner"
                )
                is_dd_val = (
                    str(row_data.get("Dry Dock", "")).strip().lower()
                    in ["yes", "true", "1"]
                )
 
                # ── Determine CREATE vs UPDATE ─────────────────────────────
                defect_id_from_row = (
                    str(row_data.get("Defect ID", "") or "").strip()
                )
                existing_defect = None
 
                if defect_id_from_row and not defect_id_from_row.startswith("#"):
                    existing_res = await db.execute(
                        select(Defect).where(
                            Defect.defect_number == defect_id_from_row,
                            Defect.is_deleted == False,
                        )
                    )
                    existing_defect = existing_res.scalars().first()
 
                # ── UPDATE PATH ────────────────────────────────────────────
                if existing_defect:
                    existing_defect.equipment_name = equip_str
                    existing_defect.title = equip_str
                    existing_defect.description = desc_str
                    existing_defect.defect_source = source_enum
                    existing_defect.priority = priority_enum
                    existing_defect.status = status_enum
                    existing_defect.date_identified = date_id
                    existing_defect.target_close_date = date_dl
                    existing_defect.is_owner = is_owner_val
                    existing_defect.is_dd = is_dd_val
                    existing_defect.updated_at = datetime.utcnow()
                    # FIX: (or 0) + 1 so None → 1, not None → 2
                    existing_defect.version = (existing_defect.version or 0) + 1
 
                    if (
                        status_enum == DefectStatus.CLOSED
                        and existing_defect.closed_at is None
                    ):
                        existing_defect.closed_at = datetime.utcnow()
                        existing_defect.closed_by_id = current_user.id
 
                    db.add(existing_defect)
                    defects_to_update.append(existing_defect)
 
                    # Handle PR entries for updated defect
                    pr_aliases = ["PR Number", "PR No", "PR No.", "PR #", "PR Details"]
                    pr_val = next(
                        (row_data[a] for a in pr_aliases if row_data.get(a)), None
                    )
                    if pr_val and str(pr_val).strip().lower() not in [
                        "none", "nan", "", "null"
                    ]:
                        for pr_no in [
                            p.strip() for p in str(pr_val).split(",") if p.strip()
                        ]:
                            existing_pr_res = await db.execute(
                                select(PrEntry).where(
                                    PrEntry.defect_id == existing_defect.id,
                                    PrEntry.pr_number == pr_no,
                                    PrEntry.is_deleted == False,
                                )
                            )
                            if not existing_pr_res.scalars().first():
                                pr_id = uuid.uuid4()
                                cache_result = await db.execute(
                                    select(MariappsPrCache).where(MariappsPrCache.requisition_no == pr_no)
                                )
                                cached = cache_result.scalars().first()
                                prs_to_insert.append(PrEntry(
                                    id=pr_id,
                                    defect_id=existing_defect.id,
                                    pr_number=pr_no,
                                    pr_description="Updated via Excel",
                                    created_by_id=current_user.id,
                                    is_deleted=False,
                                    version=1,
                                    updated_at=datetime.utcnow(),
                                    origin="VESSEL" if _should_sync() else "SHORE",
                                    mariapps_pr_status=cached.status if cached else None,
                                    created_at=datetime.utcnow(),
                                ))
 
                    # FIX: increment BEFORE the update path ends — do NOT continue yet
                    updated_count += 1
                    continue  # skip the create path below
 
                # ── CREATE PATH (no matching Defect ID found) ──────────────
                dup_res = await db.execute(
                    select(Defect).where(
                        Defect.vessel_imo == vessel.imo,
                        Defect.equipment_name == equip_str,
                        Defect.description == desc_str,
                        Defect.date_identified == date_id,
                        Defect.defect_source == source_enum,
                        Defect.is_deleted == False,
                    )
                )
                if dup_res.scalars().first():
                    raise ValueError("Duplicate: defect already exists.")
                
                from sqlalchemy import text as sa_text
                vessel_result = await control_db.execute(
                    select(Vessel).where(Vessel.imo == vessel.imo)
                )
                vessel_for_num = vessel_result.scalars().first()
                vessel_name_for_num = vessel_for_num.name if vessel_for_num else vessel.imo
                prefix = vessel_name_for_num.replace(" ", "").upper()[:6]

                seq_result = await db.execute(sa_text("""
                    INSERT INTO vessel_defect_sequences (vessel_imo, next_seq)
                    VALUES (:imo, 1)
                    ON CONFLICT (vessel_imo)
                    DO UPDATE SET next_seq = vessel_defect_sequences.next_seq + 1
                    RETURNING next_seq
                """), {"imo": vessel.imo})
                next_seq = seq_result.scalar()
                defect_number = f"{prefix}#{str(next_seq).zfill(4)}"

                new_id = uuid.uuid4()
                new_defect = Defect(
                    id=new_id,
                    defect_number=defect_number,
                    vessel_imo=vessel.imo,
                    reported_by_id=current_user.id,
                    title=equip_str,
                    equipment_name=equip_str,
                    description=desc_str,
                    defect_source=source_enum,
                    priority=priority_enum,
                    status=status_enum,
                    date_identified=date_id,
                    target_close_date=date_dl,
                    responsibility="Engine Dept",
                    pr_status="Not Set",
                    is_owner=is_owner_val,
                    is_deleted=False,
                    before_image_required=False,
                    after_image_required=False,
                    version=1,
                    origin="VESSEL" if _should_sync() else "SHORE",
                    created_at=datetime.utcnow(),
                    updated_at=datetime.utcnow(),
                    is_dd=is_dd_val,
                    closed_at=(
                        date_dl
                        if status_enum == DefectStatus.CLOSED and date_dl
                        else (date_id if status_enum == DefectStatus.CLOSED else None)
                    ),
                    closed_by_id=(
                        current_user.id if status_enum == DefectStatus.CLOSED else None
                    ),
                )
                defects_to_insert.append(new_defect)
                if _should_sync():
                    syncs_to_insert.append(
                        SyncQueue(
                            entity_id=new_id,
                            entity_type="DEFECT",
                            operation="CREATE",
                            payload={
                                "id": str(new_id),
                                "vessel_imo": vessel.imo,
                                "reported_by_id": str(current_user.id),
                                "title": equip_str,
                                "equipment_name": equip_str,
                                "description": desc_str,
                                "defect_source": source_enum.value,
                                "priority": priority_enum.value,
                                "status": status_enum.value,
                                "date_identified": (
                                    date_id.isoformat() if date_id else None
                                ),
                                "target_close_date": (
                                    date_dl.isoformat() if date_dl else None
                                ),
                                "responsibility": "Engine Dept",
                                "pr_status": "Not Set",
                                "is_owner": is_owner_val,
                                "is_dd": is_dd_val,
                                "is_deleted": False,
                                "before_image_required": False,
                                "after_image_required": False,
                            },
                            status="PENDING",
                            origin=(
                                "VESSEL"
                                if current_user.role == UserRole.VESSEL
                                else "SHORE"
                            ),
                            sync_scope="DEFECT",
                            version=1,
                        )
                    )
 
                pr_aliases = ["PR Number", "PR No", "PR No.", "PR #", "PR Details"]
                pr_val = next(
                    (row_data[a] for a in pr_aliases if row_data.get(a)), None
                )
                if pr_val and str(pr_val).strip().lower() not in [
                    "none", "nan", "", "null"
                ]:
                    for pr_no in [
                        p.strip() for p in str(pr_val).split(",") if p.strip()
                    ]:
                        pr_id = uuid.uuid4()
                        cache_result = await db.execute(
                            select(MariappsPrCache).where(MariappsPrCache.requisition_no == pr_no)
                        )
                        cached = cache_result.scalars().first()
                        prs_to_insert.append(PrEntry(
                            id=pr_id,
                            defect_id=new_id,
                            pr_number=pr_no,
                            pr_description="Imported via Excel",
                            created_by_id=current_user.id,
                            is_deleted=False,
                            version=1,
                            updated_at=datetime.utcnow(),
                            origin="VESSEL" if _should_sync() else "SHORE",
                            mariapps_pr_status=cached.status if cached else None,
                            created_at=datetime.utcnow()
                        ))
 
                        if _should_sync():
                            syncs_to_insert.append(
                                SyncQueue(
                                    entity_id=pr_id,
                                    entity_type="PR_ENTRY",
                                    operation="CREATE",
                                    payload={
                                        "id": str(pr_id),
                                        "defect_id": str(new_id),
                                        "pr_number": pr_no,
                                        "pr_description": "Imported via Excel",
                                        "created_by_id": str(current_user.id),
                                        "is_deleted": False,
                                    },
                                    status="PENDING",
                                    origin=(
                                        "VESSEL"
                                        if current_user.role == UserRole.VESSEL
                                        else "SHORE"
                                    ),
                                    sync_scope="DEFECT",
                                    version=1,
                                )
                            )
 
                created_count += 1
 
            except Exception as row_error:
                errors.append({"row": row_idx, "error": str(row_error)})
                if not skip_errors:
                    raise HTTPException(
                        status_code=400, detail=f"Row {row_idx}: {str(row_error)}"
                    )
 
        # FIX: do NOT early-return if defects_to_insert is empty —
        # there may still be updates (defects_to_update) and PRs (prs_to_insert) to commit.
        total_success = created_count + updated_count
        if total_success == 0 and not errors:
            return {
                "status": "failure",
                "message": "Import completed — no valid rows found",
                "total_rows_processed": sheet.max_row - 1,
                "created_count": 0,
                "updated_count": 0,
                "success_count": 0,
                "error_count": len(errors),
                "errors": errors,
            }
 
        try:
            # Commit new defects, updated defects (already db.add()-ed above),
            # new PR entries (for both create and update paths), and sync queue.
            if defects_to_insert:
                db.add_all(defects_to_insert)
            if prs_to_insert:
                db.add_all(prs_to_insert)
            if syncs_to_insert:
                db.add_all(syncs_to_insert)
            # Note: defects_to_update objects were already added via db.add() in the loop.
 
            await db.commit()
            logger.info(
                f"✅ Bulk import committed: {created_count} created, "
                f"{updated_count} updated"
            )
        except Exception as e:
            await db.rollback()
            logger.error(f"❌ Bulk import DB error: {e}", exc_info=True)
            raise HTTPException(
                status_code=500, detail="Database error during bulk commit."
            )
 
        # Notifications per affected vessel
        affected_vessels = list(set(
            d.vessel_imo for d in defects_to_insert + defects_to_update
        ))
        for v_imo in affected_vessels:
            v_count = sum(1 for d in defects_to_insert if d.vessel_imo == v_imo)
            try:
                vessel_result = await control_db.execute(
                    select(Vessel).where(Vessel.imo == v_imo)
                )
                vessel_obj = vessel_result.scalars().first()
                vessel_name = vessel_obj.name if vessel_obj else v_imo
                background_tasks.add_task(
                    notify_vessel_users,
                    db=db,
                    control_db=control_db,
                    vessel_imo=v_imo,
                    vessel_name=vessel_name,
                    title="Bulk Import Success",
                    message=(
                        f"{v_count} defects imported via Excel "
                        f"by {current_user.full_name}."
                    ),
                    exclude_user_id=current_user.id,
                    defect_id=None,
                )
            except Exception:
                pass
 
        # Determine overall status
        if total_success == 0:
            status_label = "failure"
        elif errors:
            status_label = "partial_success"
        else:
            status_label = "success"
 
        parts = []
        if created_count:
            parts.append(f"{created_count} created")
        if updated_count:
            parts.append(f"{updated_count} updated")
        summary = ", ".join(parts) if parts else "no changes"
 
        return {
            "status": status_label,
            "message": f"Import completed — {summary}",
            "total_rows_processed": sheet.max_row - 1,
            "created_count": created_count,
            "updated_count": updated_count,
            "success_count": total_success,        # backwards-compatible field
            "error_count": len(errors),
            "errors": errors,
        }
 
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"❌ Global import error: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# VESSEL USERS FOR DEFECT
# =============================================================================
@router.get("/{defect_id}/vessel-users", response_model=list[VesselUserResponse])
async def get_vessel_users_for_defect(
    defect_id: UUID,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
):
    """Returns vessel crew AND all shore/admin users (for @mention tagging in threads)."""
    defect = await db.get(Defect, defect_id)
    if not defect:
        raise HTTPException(status_code=404, detail="Defect not found")

    query = (
        select(User)
        .distinct()
        .outerjoin(User.vessels)
        .where(
            (Vessel.imo == defect.vessel_imo)
            | (User.role.in_([UserRole.SHORE, UserRole.ADMIN]))
        )
    )
    result = await control_db.execute(query)
    users = result.scalars().all()

    return [
        {
            "id": u.id,
            "full_name": u.full_name,
            "role": u.role.value if hasattr(u.role, "value") else str(u.role),
            "job_title": u.job_title,
        }
        for u in users
    ]


# =============================================================================
# CREATE DEFECT
# DefectService handles: DB write, SyncQueue, vessel notifications.
# This endpoint adds: vessel existence check, live feed, email background task.
# =============================================================================
@router.post("/", response_model=DefectResponse)
async def create_defect(
    defect_in: DefectCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new defect. SyncQueue + notifications handled by DefectService."""
    try:
        logger.info(
            f"🆕 Creating defect for vessel: {defect_in.vessel_imo} | {defect_in.equipment}"
        )

        vessel_result = await control_db.execute(
            select(Vessel).where(Vessel.imo == defect_in.vessel_imo)
        )
        vessel = vessel_result.scalars().first()

        if not vessel:
            raise HTTPException(status_code=404, detail="Vessel not found")

        new_defect = await DefectService.create_defect(
            db, control_db, defect_in, current_user
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error creating defect: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

    # Live feed (non-blocking, non-fatal — defect is already committed above)
    try:
        await feed_defect_opened(
            db, control_db, defect=new_defect, actor_id=current_user.id
        )
        await db.commit()
    except Exception as e:
        logger.error(f"⚠️ Live feed error (non-fatal): {e}")

    email_data = prepare_email_data(new_defect)
    background_tasks.add_task(send_defect_email, email_data, "CREATED")
    return new_defect


# =============================================================================
# GET SINGLE DEFECT
# =============================================================================
@router.get("/{defect_id}", response_model=DefectResponse)
async def get_defect(defect_id: UUID, db: AsyncSession = Depends(get_db)):
    """Get a specific defect by ID."""
    try:
        defect = await db.get(Defect, defect_id)
        if not defect or defect.is_deleted:
            raise HTTPException(status_code=404, detail="Defect not found")
        await db.refresh(defect, attribute_names=["pr_entries"])
        # Use __dict__ to avoid mutating the ORM collection (prevents SQLAlchemy
        # from nullifying defect_id on removed items at session flush)
        defect.__dict__['pr_entries'] = [pr for pr in defect.pr_entries if not pr.is_deleted]
        return defect
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error fetching defect: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# UPDATE DEFECT
# DefectService handles: DB write, full status machine, image validation,
#                        SyncQueue, notifications, priority system thread.
# This endpoint adds: live feed calls using pre/post-update state comparison.
# =============================================================================
# =============================================================================
# REPLACE your existing @router.patch("/{defect_id}") with this entire block
# =============================================================================


@router.patch("/{defect_id}", response_model=DefectResponse)
async def update_defect(
    defect_id: UUID,
    defect_update: DefectUpdate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """Update a defect. Full status machine + SyncQueue handled inside DefectService."""
    try:
        # Capture old state BEFORE delegating — needed only for live feed comparison
        old_defect = await db.get(Defect, defect_id)
        if not old_defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        old_priority = old_defect.priority
        old_status = old_defect.status
        old_before_req = old_defect.before_image_required
        old_after_req = old_defect.after_image_required
        update_data = defect_update.model_dump(exclude_unset=True)

        # ── All DB writes, status machine, SyncQueue → service ───────────────
        updated_defect = await DefectService.update_defect(
            db, control_db, defect_id, defect_update, current_user
        )

    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        import traceback

        traceback.print_exc()  # ← ADD THIS LINE
        logger.error(f"❌ Error updating defect: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    if not updated_defect:
        raise HTTPException(status_code=404, detail="Defect not found")

    # ── Live feed calls (non-blocking, non-fatal) ────────────────────────────
    try:
        new_priority = updated_defect.priority
        new_status = updated_defect.status

        if old_priority != new_priority:
            old_str = (
                old_priority.value
                if hasattr(old_priority, "value")
                else str(old_priority)
            )
            new_str = (
                new_priority.value
                if hasattr(new_priority, "value")
                else str(new_priority)
            )
            await feed_priority_changed(
                db,
                control_db,
                defect=updated_defect,
                old_priority=old_str,
                new_priority=new_str,
                actor_id=current_user.id,
            )

        if new_status == DefectStatus.CLOSED and old_status != DefectStatus.CLOSED:
            await feed_defect_closed(
                db,
                control_db,
                defect=updated_defect,
                actor_id=current_user.id,
                remarks=updated_defect.closure_remarks,
            )

        if (
            "before_image_required" in update_data
            and update_data["before_image_required"] != old_before_req
        ):
            await feed_pic_mandatory_changed(
                db,
                control_db,
                defect=updated_defect,
                image_field="before_image_required",
                is_now_required=update_data["before_image_required"],
                actor_id=current_user.id,
            )

        if (
            "after_image_required" in update_data
            and update_data["after_image_required"] != old_after_req
        ):
            await feed_pic_mandatory_changed(
                db,
                control_db,
                defect=updated_defect,
                image_field="after_image_required",
                is_now_required=update_data["after_image_required"],
                actor_id=current_user.id,
            )

        await db.commit()
    except Exception as e:
        logger.error(f"⚠️ Live feed error (non-fatal): {e}")

    # ── Background email ─────────────────────────────────────────────────────
    email_data = prepare_email_data(updated_defect)
    background_tasks.add_task(send_defect_email, email_data, "UPDATED")

    return updated_defect


# =============================================================================
# SHORE DIRECT CLOSURE — Closes immediately, SHORE/ADMIN only
# No PENDING_CLOSURE step required for shore users.
# =============================================================================
@router.patch("/{defect_id}/shore-close", response_model=DefectResponse)
async def shore_close_defect(
    defect_id: UUID,
    close_data: ShoreCloseRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Shore-side direct closure.
    Closes defect immediately to CLOSED regardless of current status.
    No PENDING_CLOSURE step required. SHORE/ADMIN only.
    Requires closure_remarks of at least 50 characters.
    """
    try:
        defect = await db.get(Defect, defect_id)
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        if defect.status == DefectStatus.CLOSED:
            raise HTTPException(status_code=400, detail="Defect is already closed")

        if current_user.role not in [UserRole.SHORE, UserRole.ADMIN]:
            raise HTTPException(
                status_code=403, detail="Only shore users can directly close a defect"
            )

        if (
            not close_data.closure_remarks
            or len(close_data.closure_remarks.strip()) < 50
        ):
            raise HTTPException(
                status_code=400,
                detail="Closure remarks must be at least 50 characters",
            )

        previous_status = (
            defect.status.value
            if hasattr(defect.status, "value")
            else str(defect.status)
        )

        defect.status = DefectStatus.CLOSED
        defect.closure_remarks = close_data.closure_remarks.strip()
        defect.closed_at = datetime.now()
        defect.closed_by_id = current_user.id

        # System thread visible to all (is_internal=False)
        db.add(
            Thread(
                id=uuid.uuid4(),
                defect_id=defect.id,
                user_id=current_user.id,
                author_role="SYSTEM",
                is_system_message=True,
                is_internal=False,
                body=f" Defect CLOSED by {current_user.full_name} (Shore). Previous status: {previous_status}",
            )
        )

        await db.commit()

        # Live feed (non-fatal)
        try:
            await feed_defect_closed(
                db,
                control_db,
                defect=defect,
                actor_id=current_user.id,
                remarks=defect.closure_remarks,
            )
            await db.commit()
        except Exception as e:
            logger.error(f"⚠️ Live feed error (non-fatal): {e}")

        await db.refresh(defect, attribute_names=["pr_entries"])

        # Notification to vessel crew
        try:
            vessel_result = await control_db.execute(
                select(Vessel).where(Vessel.imo == defect.vessel_imo)
            )
            vessel = vessel_result.scalars().first()
            vessel_name = vessel.name if vessel else defect.vessel_imo
            await notify_vessel_users(
                db=db,
                control_db=control_db,
                vessel_imo=defect.vessel_imo,
                vessel_name=vessel_name,
                title="Defect Closed",
                message=f"Defect closed by shore: {defect.title}",
                exclude_user_id=current_user.id,
                defect_id=str(defect.id),
            )
            await db.commit()
        except Exception as notif_error:
            logger.error(f"⚠️ Shore closure notification failed: {notif_error}")

        email_data = prepare_email_data(defect)
        background_tasks.add_task(send_defect_email, email_data, "CLOSED")
        return defect

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error in shore closure: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# CLOSE DEFECT (Legacy endpoint — close with evidence images)
# =============================================================================
@router.patch("/{defect_id}/close", response_model=DefectResponse)
async def close_defect(
    defect_id: UUID,
    close_data: DefectCloseRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """Close a defect with closure remarks and evidence images (legacy endpoint)."""
    try:
        closed_defect = await DefectService.close_defect(
            db, control_db, defect_id, close_data, current_user
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error closing defect: {str(e)}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    if not closed_defect:
        raise HTTPException(status_code=404, detail="Defect not found")

    try:
        await feed_defect_closed(
            db,
            control_db,
            defect=closed_defect,
            actor_id=current_user.id,
            remarks=closed_defect.closure_remarks,
        )
        await db.commit()
    except Exception as e:
        logger.error(f"⚠️ Live feed error (non-fatal): {e}")

    email_data = prepare_email_data(closed_defect)
    background_tasks.add_task(send_defect_email, email_data, "CLOSED")
    return closed_defect


# =============================================================================
# DELETE DEFECT (Soft)
# =============================================================================
@router.delete("/{defect_id}")
async def remove_defect(
    defect_id: UUID,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft delete a defect (is_deleted=True). Sends removal email."""
    try:
        defect = await db.get(Defect, defect_id)
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        email_data = prepare_email_data(defect)
        result = await DefectService.delete_defect(db, defect_id)
        if not result:
            raise HTTPException(status_code=404, detail="Defect not found")

        background_tasks.add_task(send_defect_email, email_data, "REMOVED")
        return {"message": "Defect removed and archived"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error removing defect: {str(e)}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# THREADS
# Flat POST /threads URL — matches original production routes.
# is_internal support with SHORE/ADMIN role guard.
# Smart @mention filtering — vessel users excluded from internal message notifications.
# Idempotent — returns existing thread if UUID already present.
# =============================================================================
@router.post("/threads", response_model=ThreadResponse)
async def create_thread(
    thread_in: ThreadCreate,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Create a thread/comment on a defect.

    is_internal=True  → SHORE/ADMIN only. Vessel users cannot create or see these.
    is_internal=False → Visible to all users (default).

    Idempotent: if a thread with the same UUID already exists, returns it.

    @mention notifications: vessel users are excluded from internal message notifications.
    """
    try:
        # Idempotency check
        existing = await db.get(Thread, thread_in.id)
        if existing:
            res = await db.execute(
                select(Thread)
                .where(Thread.id == thread_in.id)
                .options(selectinload(Thread.attachments))
            )
            return res.scalars().first()

        if thread_in.is_internal and current_user.role not in [
            UserRole.SHORE,
            UserRole.ADMIN,
        ]:
            logger.warning(
                f"⚠️ User {current_user.id} ({current_user.role}) tried to create internal message"
            )
            raise HTTPException(
                status_code=403,
                detail="Only shore users can create internal messages",
            )

        new_thread = Thread(
            id=thread_in.id,
            defect_id=thread_in.defect_id,
            user_id=current_user.id,
            author_role=thread_in.author,
            body=thread_in.body,
            tagged_user_ids=thread_in.tagged_user_ids,
            is_internal=thread_in.is_internal,
        )
        db.add(new_thread)

        # Smart mention notifications — vessel users excluded from internal messages
        if thread_in.tagged_user_ids:
            defect = await db.get(Defect, thread_in.defect_id)
            eligible_user_ids = []
            for user_id in thread_in.tagged_user_ids:
                user_result = await control_db.execute(
                    select(User).where(User.id == user_id)
                )
                tagged_user = user_result.scalar_one_or_none()
                if tagged_user:
                    should_notify = not thread_in.is_internal or tagged_user.role in [
                        UserRole.SHORE,
                        UserRole.ADMIN,
                    ]
                    if should_notify:
                        eligible_user_ids.append(user_id)
                    else:
                        logger.info(
                            f"⏭️ Skipping vessel user {tagged_user.full_name} for internal message"
                        )

            if eligible_user_ids:
                await create_task_for_mentions(
                    db=db,
                    control_db=control_db,
                    defect_id=thread_in.defect_id,
                    defect_title=defect.title if defect else "Defect",
                    creator_id=current_user.id,
                    tagged_user_ids=eligible_user_ids,
                    comment_body=thread_in.body,
                    is_internal=thread_in.is_internal,
                    thread_id=str(new_thread.id),
                )
                await feed_mention(
                    db,
                    control_db,
                    defect=defect,
                    thread_body=thread_in.body,
                    mentioned_user_ids=eligible_user_ids,
                    actor_id=current_user.id,
                    is_internal=thread_in.is_internal,
                )

        await db.commit()
        await db.refresh(new_thread, attribute_names=["attachments"])
        logger.info(
            f"{'🔒 Internal' if thread_in.is_internal else '🌐 External'} thread created: {new_thread.id}"
        )
        return new_thread

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error creating thread: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{defect_id}/threads", response_model=list[ThreadResponse])
async def get_defect_threads(
    defect_id: UUID,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Get all threads for a defect.

    SHORE/ADMIN → sees ALL messages (internal + external).
    VESSEL      → sees ONLY external messages (is_internal=False).
    """
    try:
        query = (
            select(Thread)
            .where(Thread.defect_id == defect_id)
            .options(selectinload(Thread.attachments))
        )

        if current_user.role not in [UserRole.SHORE, UserRole.ADMIN]:
            query = query.where(Thread.is_internal == False)
            logger.info(
                f"🚢 Filtering internal messages for vessel user: {current_user.full_name}"
            )
        else:
            logger.info(f"👔 Shore user {current_user.full_name} can see all messages")

        query = query.order_by(Thread.created_at.asc())
        result = await db.execute(query)
        threads = result.scalars().all()
        for thread in threads:
            if thread.is_system_message:
                thread.author_role = "SYSTEM"
            else:
                user_result = await control_db.execute(
                    select(User).where(User.id == thread.user_id)
                )
                user = user_result.scalars().first()
                thread.author_role = (
                    user.full_name if user else (thread.author_role or "Unknown")
                )

        logger.info(f"✅ Returning {len(threads)} threads to {current_user.role} user")
        return threads

    except Exception as e:
        logger.error(f"❌ Error fetching threads: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# ATTACHMENTS — Flat URL, 1MB limit, idempotent
# =============================================================================
@router.post("/attachments", response_model=AttachmentResponse)
async def create_attachment(
    attachment_in: AttachmentBase,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Save attachment metadata. 1MB file size limit.
    Idempotent — returns existing attachment if UUID already saved.
    The blob must already be uploaded to Azure by the browser via a write SAS URL.
    """
    try:
        existing = await db.get(Attachment, attachment_in.id)
        if existing:
            logger.info(
                f"⚠️ Attachment {attachment_in.id} already exists, returning existing"
            )
            return existing

        MAX_FILE_SIZE = 1024 * 1024  # 1MB
        if attachment_in.file_size and attachment_in.file_size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"File '{attachment_in.file_name}' exceeds 1MB limit "
                    f"({attachment_in.file_size / 1024 / 1024:.2f}MB)"
                ),
            )

        new_attachment = Attachment(
            id=attachment_in.id,
            thread_id=attachment_in.thread_id,
            file_name=attachment_in.file_name,
            file_size=attachment_in.file_size,
            content_type=attachment_in.content_type,
            blob_path=attachment_in.blob_path,
            origin="SHORE",  # ← ADD THIS
            version=1,  # ← ADD THIS
            updated_at=datetime.utcnow(),  # ← ADD THIS
        )
        db.add(new_attachment)
        await db.commit()
        await db.refresh(new_attachment)
        logger.info(f"✅ Attachment created: {new_attachment.file_name}")
        return new_attachment

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error creating attachment: {str(e)}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{defect_id}/threads/{thread_id}/attachments/{attachment_id}/url")
async def get_attachment_url(
    defect_id: UUID,
    thread_id: UUID,
    attachment_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a SAS download URL for a specific attachment."""
    try:
        attachment = await db.get(Attachment, attachment_id)
        if not attachment or attachment.thread_id != thread_id:
            raise HTTPException(status_code=404, detail="Attachment not found")
        return {"url": generate_read_sas_url(attachment.blob_path)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# IMAGES
# Flat POST /images (defect_id from body — matches original production routes).
# GET /{defect_id}/images/{image_type} — path-based type filter.
# DELETE /{defect_id}/images/{image_id} — hard delete via DefectService.
# =============================================================================
@router.post("/images", response_model=DefectImageResponse)
async def save_defect_image(
    image_data: DefectImageCreate,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """
    Save before/after image metadata.
    Idempotent — returns existing image if UUID already saved.
    Calls feed_image_uploaded after successful save.
    SyncQueue written by DefectService on vessel mode.
    """
    try:
        logger.info(
            f"💾 Saving {image_data.image_type} image for defect {image_data.defect_id}"
        )

        # Idempotency check
        existing = await db.get(DefectImage, image_data.id)
        if existing:
            logger.info(f"⚠️ Image {image_data.id} already exists, returning existing")
            return DefectImageResponse(
                id=existing.id,
                defect_id=existing.defect_id,
                image_type=existing.image_type,
                file_name=existing.file_name,
                file_size=existing.file_size,
                blob_path=existing.blob_path,
                created_at=existing.created_at,
                image_url=generate_read_sas_url(existing.blob_path),
            )

        new_image = await DefectService.save_defect_image(
            db, image_data.defect_id, image_data
        )

        # Live feed (non-fatal)
        defect = await db.get(Defect, image_data.defect_id)
        if defect:
            try:
                await feed_image_uploaded(
                    db,
                    control_db,
                    defect=defect,
                    image_type=image_data.image_type,
                    file_name=image_data.file_name,
                    actor_id=current_user.id,
                )
                await db.commit()
            except Exception as e:
                logger.error(f"⚠️ Live feed error (non-fatal): {e}")

        logger.info(f"✅ Image saved: {new_image.file_name}")
        return DefectImageResponse(
            id=new_image.id,
            defect_id=new_image.defect_id,
            image_type=new_image.image_type,
            file_name=new_image.file_name,
            file_size=new_image.file_size,
            blob_path=new_image.blob_path,
            created_at=new_image.created_at,
            image_url=generate_read_sas_url(new_image.blob_path),
        )

    except Exception as e:
        logger.error(f"❌ Error saving image: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get(
    "/{defect_id}/images/{image_type}", response_model=list[DefectImageResponse]
)
async def get_defect_images(
    defect_id: UUID,
    image_type: str,
    db: AsyncSession = Depends(get_db),
):
    """Get all images of a specific type (before/after) for a defect."""
    try:
        logger.info(f"📷 Fetching {image_type} images for defect {defect_id}")
        if image_type not in ["before", "after"]:
            raise HTTPException(
                status_code=400, detail="image_type must be 'before' or 'after'"
            )

        result = await db.execute(
            select(DefectImage)
            .where(
                DefectImage.defect_id == defect_id,
                DefectImage.image_type == image_type,
            )
            .order_by(DefectImage.created_at.asc())
        )
        images = result.scalars().all()
        logger.info(f"✅ Found {len(images)} {image_type} images")

        return [
            DefectImageResponse(
                id=img.id,
                defect_id=img.defect_id,
                image_type=img.image_type,
                file_name=img.file_name,
                file_size=img.file_size,
                blob_path=img.blob_path,
                created_at=img.created_at,
                image_url=generate_read_sas_url(img.blob_path),
            )
            for img in images
        ]

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error fetching images: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{defect_id}/images/{image_id}")
async def delete_defect_image(
    defect_id: UUID,
    image_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a defect image. SyncQueue entry written by DefectService on vessel."""
    try:
        deleted = await DefectService.delete_defect_image(db, defect_id, image_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Image not found")
        return {"message": "Image deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error deleting image: {str(e)}")
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# PR ENTRIES — Flat URLs matching original production routes
# =============================================================================
import re
PR_FORMAT_REGEX = re.compile(r'^[A-Z]{2,5}\/(V|O)-\d{4}\/REQ\d{2}$')

@router.post("/pr-entries", response_model=PrEntryResponse)
async def create_pr_entry(
    pr_in: PrEntryCreate,
    db: AsyncSession = Depends(get_db),
    control_db: AsyncSession = Depends(get_control_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new PR entry for a defect. Fires feed_pr_added live feed event."""
    try:
        logger.info(
            f"📝 Creating PR entry for defect {pr_in.defect_id}: {pr_in.pr_number}"
        )

        defect = await db.get(Defect, pr_in.defect_id)
        if not defect:
            raise HTTPException(status_code=404, detail="Defect not found")

        new_pr = await DefectService.create_pr_entry(
            db, pr_in.defect_id, pr_in, current_user
        )

        # Live feed (non-fatal)
        try:
            await feed_pr_added(db, control_db, defect=defect, pr_number=pr_in.pr_number, actor_id=current_user.id)
            if not PR_FORMAT_REGEX.match(pr_in.pr_number):
                await feed_pr_invalid_format(db, control_db, defect=defect, pr_number=pr_in.pr_number, actor_id=current_user.id)
            await db.commit()
        except Exception as e:
            logger.error(f"⚠️ Live feed error (non-fatal): {e}")

        logger.info(f"✅ PR entry created: {new_pr.id}")
        return new_pr

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Error creating PR entry: {str(e)}", exc_info=True)
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/pr-entries/{pr_id}", response_model=PrEntryResponse)
async def update_pr_entry(
    pr_id: UUID,
    pr_in: PrEntryUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update PR entry number and/or description."""
    try:
        pr_entry = await db.get(PrEntry, pr_id)
        if not pr_entry or pr_entry.is_deleted:
            raise HTTPException(status_code=404, detail="PR entry not found")

        result = await DefectService.update_pr_entry(
            db, pr_entry.defect_id, pr_id, pr_in
        )
        if not result:
            raise HTTPException(status_code=404, detail="PR entry not found")

        logger.info(f"✏️ PR entry updated: {pr_id}")
        return result

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"❌ Error updating PR entry: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to update PR entry")


@router.get("/{defect_id}/pr-entries", response_model=list[PrEntryResponse])
async def get_pr_entries(defect_id: UUID, db: AsyncSession = Depends(get_db)):
    """Get all active (non-deleted) PR entries for a defect."""
    try:
        result = await db.execute(
            select(PrEntry)
            .where(PrEntry.defect_id == defect_id, PrEntry.is_deleted == False)
            .order_by(PrEntry.created_at.asc())
        )
        return result.scalars().all()
    except Exception as e:
        logger.error(f"❌ Error fetching PR entries: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/pr-entries/{pr_id}")
async def delete_pr_entry(
    pr_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft delete a PR entry."""
    try:
        pr_entry = await db.get(PrEntry, pr_id)
        if not pr_entry or pr_entry.is_deleted:
            raise HTTPException(status_code=404, detail="PR entry not found")

        deleted = await DefectService.delete_pr_entry(db, pr_entry.defect_id, pr_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="PR entry not found")

        logger.info(f"🗑️ PR entry soft-deleted: {pr_id}")
        return {"message": "PR entry deleted", "pr_id": pr_id}

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.error(f"❌ Error deleting PR entry: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to delete PR entry")


# =============================================================================
# SYSTEM THREAD HELPER
# =============================================================================
async def create_system_thread(
    db: AsyncSession,
    defect_id: UUID,
    message: str,
    is_internal: bool = False,
    user_id: UUID = None,
):
    """
    Creates a system-generated thread message on a defect.

    is_internal=True  → Shore-only notification. Vessel users cannot see it.
    is_internal=False → Visible to all users.

    Used by other services/tasks to post automated status messages.
    Uses db.flush() (not commit) so the caller controls the transaction boundary.
    """
    system_thread = Thread(
        id=uuid.uuid4(),
        defect_id=defect_id,
        user_id=user_id,
        author_role="SYSTEM",
        body=message,
        is_system_message=True,
        is_internal=is_internal,
    )
    db.add(system_thread)
    await db.flush()
    return system_thread


@router.post("/{defect_id}/flag")
async def toggle_flag(
    defect_id: UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(UserDefectFlag).where(
            UserDefectFlag.user_id   == current_user.id,
            UserDefectFlag.defect_id == defect_id,
        )
    )
    row = result.scalar_one_or_none()

    if row:
        await db.delete(row)
        flagged = False
    else:
        db.add(UserDefectFlag(user_id=current_user.id, defect_id=defect_id))
        flagged = True

    await db.commit()
    return {"defect_id": str(defect_id), "is_flagged": flagged}