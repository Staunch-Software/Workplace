# routers/export.py
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from db.database import get_db
from models.schema import Ticket
from core.deps import require_role
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("")
async def export_excel(
    vesselName: str = Query(None),
    user=Depends(require_role("SHORE", "ADMIN")),
    db: AsyncSession = Depends(get_db),
):
    q = select(Ticket).where(Ticket.vesselName.isnot(None))
    if vesselName and vesselName != "all":
        q = q.where(Ticket.vesselName == vesselName)
    q = q.order_by(Ticket.jiraSortOrder.asc().nulls_last())
    tickets = (await db.execute(q)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["Reference","Summary","Vessel","Priority","Status","Module","Environment","Requester","Created"]
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row, t in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=t.reference or "PENDING")
        ws.cell(row=row, column=2, value=t.summary)
        ws.cell(row=row, column=3, value=t.vesselName)
        ws.cell(row=row, column=4, value=t.priority)
        ws.cell(row=row, column=5, value=t.jiraStatus or t.status)
        ws.cell(row=row, column=6, value=t.module)
        ws.cell(row=row, column=7, value=t.environment)
        ws.cell(row=row, column=8, value=t.requester)
        ws.cell(row=row, column=9, value=str(t.createdAt or ""))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ozellar_tickets.xlsx"},
    )
