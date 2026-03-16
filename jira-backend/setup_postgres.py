"""
setup_postgres.py
Run this from your backend folder:
    python setup_postgres.py

It will:
  1. Write all updated Python files (config, database, schema, routers, automation)
  2. Update .env
  3. Update main.py
  4. Write migrate_mongo_to_pg.py (run once after server starts)
  5. pip install sqlalchemy asyncpg
  6. Create the PostgreSQL database
"""
import os
import sys
import subprocess

BASE = os.path.dirname(os.path.abspath(__file__))


def write(rel_path, content):
    full = os.path.join(BASE, rel_path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with open(full, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  wrote  {rel_path}")


print()
print("=" * 55)
print("  Ozellar Portal: MongoDB -> PostgreSQL Migration")
print("=" * 55)
print()

# ─── .env ────────────────────────────────────────────────────────────────────
print("Step 1: Writing .env")
write(".env", """\
# Database (PostgreSQL)
DB_USER=Deepa
DB_PASSWORD=Admin@123
DB_HOST=localhost
DB_PORT=5432
DB_NAME=ozellar

# Security
JWT_SECRET=change_this_to_a_long_random_secret_key
JWT_EXPIRE_MINUTES=480

# Jira
JIRA_BASE_URL=https://mariapps.atlassian.net
JIRA_EMAIL=your_jira_email@mariapps.com
JIRA_PASSWORD=your_jira_password
JIRA_PROJECT_KEY=OZLR

# Playwright paths (Windows)
JIRA_COOKIES_PATH=C:/tmp/jira-cookies.json
JIRA_SCREENSHOT_DIR=C:/tmp
""")

# ─── core/config.py ──────────────────────────────────────────────────────────
print("Step 2: Writing core/config.py")
write("core/config.py", """\
# core/config.py
from pydantic_settings import BaseSettings
from urllib.parse import quote_plus


class Settings(BaseSettings):
    DB_USER: str
    DB_PASSWORD: str
    DB_HOST: str = "localhost"
    DB_PORT: str = "5432"
    DB_NAME: str

    JWT_SECRET: str
    JWT_EXPIRE_MINUTES: int = 480

    JIRA_BASE_URL: str = "https://mariapps.atlassian.net"
    JIRA_EMAIL: str = ""
    JIRA_PASSWORD: str = ""
    JIRA_PROJECT_KEY: str = "OZLR"

    @property
    def SQLALCHEMY_DATABASE_URI(self) -> str:
        encoded_password = quote_plus(self.DB_PASSWORD)
        return (
            f"postgresql+asyncpg://{self.DB_USER}:{encoded_password}"
            f"@{self.DB_HOST}:{self.DB_PORT}/{self.DB_NAME}"
        )

    class Config:
        env_file = ".env"
        case_sensitive = True
        extra = "ignore"


settings = Settings()
""")

# ─── db/database.py ──────────────────────────────────────────────────────────
print("Step 3: Writing db/database.py")
write("db/database.py", """\
# db/database.py  (replaces db/mongodb.py)
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker, declarative_base
from core.config import settings

engine = create_async_engine(
    settings.SQLALCHEMY_DATABASE_URI,
    echo=False,
    future=True,
    pool_size=10,
    max_overflow=20,
)

SessionLocal = sessionmaker(
    bind=engine,
    class_=AsyncSession,
    expire_on_commit=False,
    autoflush=False,
)

Base = declarative_base()


async def get_db():
    async with SessionLocal() as session:
        try:
            yield session
            await session.commit()
        except Exception:
            await session.rollback()
            raise
        finally:
            await session.close()


async def init_models():
    from models.schema import User, Vessel, Ticket, Comment  # noqa
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    print("Database tables ready.")
""")

# ─── models/schema.py ────────────────────────────────────────────────────────
print("Step 4: Writing models/schema.py")
write("models/schema.py", """\
# models/schema.py
import uuid
from datetime import datetime
from sqlalchemy import Column, String, Boolean, DateTime, Integer, Text, ForeignKey, Index
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.orm import relationship
from db.database import Base


def _uuid():
    return str(uuid.uuid4())


class User(Base):
    __tablename__ = "users"
    id         = Column(String, primary_key=True, default=_uuid)
    name       = Column(String, nullable=False)
    email      = Column(String, unique=True, nullable=False, index=True)
    password   = Column(String, nullable=False)
    role       = Column(String, nullable=False)
    vesselName = Column(String, nullable=True)
    createdAt  = Column(DateTime, default=datetime.utcnow)
    updatedAt  = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Vessel(Base):
    __tablename__ = "vessels"
    id        = Column(String, primary_key=True, default=_uuid)
    name      = Column(String, unique=True, nullable=False)
    code      = Column(String, unique=True, nullable=False)
    isActive  = Column(Boolean, default=True, nullable=False)
    createdAt = Column(DateTime, default=datetime.utcnow)
    updatedAt = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Ticket(Base):
    __tablename__ = "tickets"
    id                   = Column(String, primary_key=True, default=_uuid)
    reference            = Column(String, nullable=True, unique=True, index=True)
    referenceNum         = Column(Integer, nullable=True)
    summary              = Column(String, nullable=False)
    description          = Column(Text, default="")
    module               = Column(String, nullable=True)
    environment          = Column(String, nullable=True)
    priority             = Column(String, nullable=True)
    requestType          = Column(String, nullable=True)
    status               = Column(String, default="SUP IN PROGRESS")
    jiraStatus           = Column(String, nullable=True, index=True)
    jiraSubmissionStatus = Column(String, default="PENDING")
    jiraUrl              = Column(String, nullable=True)
    jiraSortOrder        = Column(Integer, nullable=True)
    vesselName           = Column(String, nullable=True, index=True)
    requester            = Column(String, nullable=True)
    createdAt            = Column(DateTime, default=datetime.utcnow)
    updatedAt            = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    jiraCreatedAt        = Column(DateTime, nullable=True)
    jiraUpdatedAt        = Column(DateTime, nullable=True)
    lastSyncedAt         = Column(DateTime, nullable=True)
    detailFetchedAt      = Column(DateTime, nullable=True)
    attachments          = Column(JSONB, default=list)
    sharedWith           = Column(JSONB, default=list)
    comments             = relationship(
        "Comment",
        back_populates="ticket",
        cascade="all, delete-orphan",
        order_by="Comment.createdAt",
    )
    __table_args__ = (
        Index("ix_tickets_updated", "updatedAt"),
        Index("ix_tickets_vessel_status", "vesselName", "jiraStatus"),
    )


class Comment(Base):
    __tablename__ = "comments"
    id        = Column(String, primary_key=True, default=_uuid)
    ticket_id = Column(String, ForeignKey("tickets.id", ondelete="CASCADE"), nullable=False, index=True)
    author    = Column(String, nullable=True)
    message   = Column(Text, nullable=False)
    source    = Column(String, default="jira")
    createdAt = Column(DateTime, default=datetime.utcnow)
    images    = Column(JSONB, default=list)
    ticket    = relationship("Ticket", back_populates="comments")
""")

# ─── main.py ─────────────────────────────────────────────────────────────────
print("Step 5: Writing main.py")
write("main.py", """\
import sys
import asyncio

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager
from db.database import init_models
from routers import auth, tickets, vessels, jira, export


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_models()
    yield


app = FastAPI(title="Ozellar MA Ticketing Portal API", version="1.0.0", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth.router)
app.include_router(tickets.router)
app.include_router(vessels.router)
app.include_router(jira.router)
app.include_router(export.router)


@app.get("/")
def root():
    return {"message": "Ozellar MA Ticketing Portal API", "status": "running"}
""")

# ─── routers/tickets.py ──────────────────────────────────────────────────────
print("Step 6: Writing routers/tickets.py")
write("routers/tickets.py", """\
# routers/tickets.py
from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_, not_
from sqlalchemy.orm import selectinload
from db.database import get_db
from models.schema import Ticket, Comment
from models.ticket import TicketCreate
from core.deps import get_current_user
from pydantic import BaseModel
from datetime import datetime
import uuid

router = APIRouter(prefix="/api/tickets", tags=["tickets"])

CLOSED_STATUSES = ["Cancelled", "Closed", "Resolved", "Resolved Awaiting Confirmation"]
VALID_STATUSES = [
    "Sup In Progress", "Dev In Progress", "In Progress",
    "Waiting for Customer", "Pending", "FSD TO REVIEW",
    "FSD APPROVED", "READY FOR UAT", "UAT IN PROGRESS",
    "Resolved Awaiting Confirmation", "Resolved", "Cancelled", "Closed",
]


def _fmt_comment(c: Comment) -> dict:
    return {
        "id":        c.id,
        "author":    c.author,
        "message":   c.message,
        "source":    c.source,
        "createdAt": c.createdAt.isoformat() if c.createdAt else None,
        "images":    c.images or [],
    }


def fmt(t: Ticket) -> dict:
    return {
        "id":                   t.id,
        "reference":            t.reference,
        "referenceNum":         t.referenceNum,
        "summary":              t.summary,
        "description":          t.description,
        "module":               t.module,
        "environment":          t.environment,
        "priority":             t.priority,
        "requestType":          t.requestType,
        "status":               t.status,
        "jiraStatus":           t.jiraStatus,
        "jiraSubmissionStatus": t.jiraSubmissionStatus,
        "jiraUrl":              t.jiraUrl,
        "jiraSortOrder":        t.jiraSortOrder,
        "vesselName":           t.vesselName,
        "requester":            t.requester,
        "attachments":          t.attachments or [],
        "sharedWith":           t.sharedWith or [],
        "comments":             [_fmt_comment(c) for c in (t.comments or [])],
        "createdAt":            t.createdAt.isoformat() if t.createdAt else None,
        "updatedAt":            t.updatedAt.isoformat() if t.updatedAt else None,
        "jiraCreatedAt":        t.jiraCreatedAt.isoformat() if t.jiraCreatedAt else None,
        "jiraUpdatedAt":        t.jiraUpdatedAt.isoformat() if t.jiraUpdatedAt else None,
        "lastSyncedAt":         t.lastSyncedAt.isoformat() if t.lastSyncedAt else None,
    }


class StatusUpdate(BaseModel):
    status: str

class CommentCreate(BaseModel):
    message: str


@router.get("")
async def get_tickets(
    vesselName:  str = Query(None),
    status:      str = Query(None),
    statusList:  str = Query(None),
    priority:    str = Query(None),
    search:      str = Query(None),
    sortBy:      str = Query("updatedAt"),
    sortOrder:   str = Query("desc"),
    page:        int = 1,
    limit:       int = 15,
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(Ticket)

    if user["role"] == "vessel":
        q = q.where(Ticket.vesselName == user["vesselName"])
    elif vesselName and vesselName != "all":
        q = q.where(Ticket.vesselName == vesselName)

    if statusList:
        statuses = [s.strip() for s in statusList.split(",") if s.strip()]
        if statuses:
            q = q.where(Ticket.jiraStatus.in_(statuses))
    elif status == "open":
        q = q.where(not_(Ticket.jiraStatus.in_(CLOSED_STATUSES)))
    elif status == "closed":
        q = q.where(Ticket.jiraStatus.in_(CLOSED_STATUSES))
    elif status and status != "all":
        q = q.where(Ticket.jiraStatus == status)

    if priority and priority != "all":
        q = q.where(Ticket.priority == priority)

    if search:
        q = q.where(or_(
            Ticket.summary.ilike(f"%{search}%"),
            Ticket.reference.ilike(f"%{search}%"),
        ))

    total = (await db.execute(select(func.count()).select_from(q.subquery()))).scalar_one()

    sort_map = {
        "reference":     Ticket.reference,
        "requester":     Ticket.requester,
        "createdAt":     Ticket.createdAt,
        "updatedAt":     Ticket.updatedAt,
        "priority":      Ticket.priority,
        "jiraSortOrder": Ticket.jiraSortOrder,
    }
    col = sort_map.get(sortBy, Ticket.updatedAt)
    q = q.order_by(col.asc().nulls_last() if sortOrder == "asc" else col.desc().nulls_last())
    q = q.options(selectinload(Ticket.comments)).offset((page - 1) * limit).limit(limit)

    tickets_raw = (await db.execute(q)).scalars().all()
    return {
        "tickets": [fmt(t) for t in tickets_raw],
        "pagination": {"page": page, "limit": limit, "total": total, "totalPages": -(-total // limit)},
    }


@router.post("")
async def create_ticket(
    body: TicketCreate,
    user=Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.utcnow()
    ticket = Ticket(
        id=str(uuid.uuid4()), summary=body.summary, description=body.description,
        module=body.module, environment=body.environment, priority=body.priority,
        vesselName=user["vesselName"], requester=user["email"],
        jiraSubmissionStatus="PENDING", status="SUP IN PROGRESS",
        attachments=[], sharedWith=[], createdAt=now, updatedAt=now,
    )
    db.add(ticket)
    await db.flush()
    await db.refresh(ticket, ["comments"])
    return fmt(ticket)


@router.get("/{ticket_id}")
async def get_ticket(ticket_id: str, user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Ticket).options(selectinload(Ticket.comments)).where(Ticket.id == ticket_id)
    )
    t = result.scalar_one_or_none()
    if not t:
        raise HTTPException(404, "Ticket not found")
    return fmt(t)


@router.post("/{ticket_id}/comments")
async def add_comment(
    ticket_id: str, body: CommentCreate,
    user=Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if not body.message.strip():
        raise HTTPException(400, "Comment message cannot be empty")
    result = await db.execute(
        select(Ticket).options(selectinload(Ticket.comments)).where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(404, "Ticket not found")
    comment = Comment(
        id=str(uuid.uuid4()), ticket_id=ticket_id,
        author=user.get("name") or user.get("email") or "User",
        message=body.message.strip(), source="portal",
        createdAt=datetime.utcnow(), images=[],
    )
    db.add(comment)
    ticket.updatedAt = datetime.utcnow()
    await db.flush()
    await db.refresh(ticket, ["comments"])
    return fmt(ticket)


@router.patch("/{ticket_id}/status")
async def update_ticket_status(
    ticket_id: str, body: StatusUpdate,
    user=Depends(get_current_user), db: AsyncSession = Depends(get_db),
):
    if body.status not in VALID_STATUSES:
        raise HTTPException(400, f"Invalid status: {body.status}")
    result = await db.execute(
        select(Ticket).options(selectinload(Ticket.comments)).where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(404, "Ticket not found")
    ticket.jiraStatus = body.status
    ticket.status     = body.status
    ticket.updatedAt  = datetime.utcnow()
    return fmt(ticket)
""")

# ─── routers/vessels.py ──────────────────────────────────────────────────────
print("Step 7: Writing routers/vessels.py")
write("routers/vessels.py", """\
# routers/vessels.py
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from db.database import get_db
from models.schema import Vessel
from models.vessel import VesselCreate
from core.deps import get_current_user, require_role
from datetime import datetime
import uuid

router = APIRouter(prefix="/api/vessels", tags=["vessels"])


def fmt(v: Vessel) -> dict:
    return {
        "id": v.id, "name": v.name, "code": v.code, "isActive": v.isActive,
        "createdAt": v.createdAt.isoformat() if v.createdAt else None,
        "updatedAt": v.updatedAt.isoformat() if v.updatedAt else None,
    }


@router.get("")
async def get_vessels(user=Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vessel).where(Vessel.isActive == True))
    return [fmt(v) for v in result.scalars().all()]


@router.post("")
async def create_vessel(body: VesselCreate, user=Depends(require_role("admin")), db: AsyncSession = Depends(get_db)):
    now = datetime.utcnow()
    vessel = Vessel(id=str(uuid.uuid4()), name=body.name, code=body.code, isActive=True, createdAt=now, updatedAt=now)
    db.add(vessel)
    await db.flush()
    return fmt(vessel)


@router.patch("/{vessel_id}")
async def update_vessel(vessel_id: str, body: dict, user=Depends(require_role("admin")), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vessel).where(Vessel.id == vessel_id))
    vessel = result.scalar_one_or_none()
    if not vessel:
        raise HTTPException(404, "Vessel not found")
    allowed = {"name", "code", "isActive"}
    updated = False
    for k, v in body.items():
        if k in allowed:
            setattr(vessel, k, v)
            updated = True
    if not updated:
        raise HTTPException(400, "No valid fields to update")
    vessel.updatedAt = datetime.utcnow()
    return fmt(vessel)


@router.delete("/{vessel_id}")
async def deactivate_vessel(vessel_id: str, user=Depends(require_role("admin")), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Vessel).where(Vessel.id == vessel_id))
    vessel = result.scalar_one_or_none()
    if not vessel:
        raise HTTPException(404, "Vessel not found")
    vessel.isActive = False
    vessel.updatedAt = datetime.utcnow()
    return {"success": True, "id": vessel_id}
""")

# ─── routers/export.py ───────────────────────────────────────────────────────
print("Step 8: Writing routers/export.py")
write("routers/export.py", """\
# routers/export.py
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from db.database import get_db
from models.schema import Ticket
from core.deps import require_role
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("")
async def export_excel(
    vesselName: str = Query(None),
    user=Depends(require_role("shore", "admin")),
    db: AsyncSession = Depends(get_db),
):
    q = select(Ticket).where(Ticket.vesselName.isnot(None))
    if vesselName and vesselName != "all":
        q = q.where(Ticket.vesselName == vesselName)
    q = q.order_by(Ticket.jiraSortOrder.asc().nulls_last())
    tickets = (await db.execute(q)).scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["Reference","Summary","Vessel","Priority","Status","Module","Environment","Requester","Created"]
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row, t in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=t.reference or "PENDING")
        ws.cell(row=row, column=2, value=t.summary)
        ws.cell(row=row, column=3, value=t.vesselName)
        ws.cell(row=row, column=4, value=t.priority)
        ws.cell(row=row, column=5, value=t.jiraStatus or t.status)
        ws.cell(row=row, column=6, value=t.module)
        ws.cell(row=row, column=7, value=t.environment)
        ws.cell(row=row, column=8, value=t.requester)
        ws.cell(row=row, column=9, value=str(t.createdAt or ""))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ozellar_tickets.xlsx"},
    )
""")

# ─── automation/pull_service.py ──────────────────────────────────────────────
print("Step 9: Writing automation/pull_service.py")
write("automation/pull_service.py", """\
# automation/pull_service.py
import asyncio
import re
import uuid
from datetime import datetime, timedelta
from sqlalchemy import select, delete
from sqlalchemy.orm import selectinload
from db.database import SessionLocal
from models.schema import Ticket, Comment
from automation.playwright_service import get_jira_service
from automation.status_map import detect_vessel_from_text, extract_reference_num


def _parse_jira_date(date_str: str) -> datetime | None:
    if not date_str:
        return None
    s = date_str.strip()
    IST_OFFSET = timedelta(hours=5, minutes=30)
    now_utc = datetime.utcnow()
    now_ist = now_utc + IST_OFFSET

    def ist_to_utc(dt): return dt - IST_OFFSET

    if s.lower() == "today":
        return ist_to_utc(datetime(now_ist.year, now_ist.month, now_ist.day, 12, 0))
    if s.lower() == "yesterday":
        return ist_to_utc(datetime(now_ist.year, now_ist.month, now_ist.day, 12, 0) - timedelta(days=1))

    m = re.match(r'^(Today|Yesterday)\\s+(\\d{1,2}):(\\d{2})\\s*(AM|PM)$', s, re.IGNORECASE)
    if m:
        label = m.group(1).lower()
        h, mi = int(m.group(2)), int(m.group(3))
        ap = m.group(4).upper()
        if ap == "PM" and h < 12: h += 12
        if ap == "AM" and h == 12: h = 0
        base = datetime(now_ist.year, now_ist.month, now_ist.day)
        if label == "yesterday": base -= timedelta(days=1)
        return ist_to_utc(base.replace(hour=h, minute=mi))

    m = re.match(r"^(\\d{1,2})[\\s/]+([A-Za-z]+)[\\s/]+(\\d{2,4})$", s)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        mn = months.get(m.group(2).lower()[:3]); yr = int(m.group(3))
        if yr < 100: yr += 2000
        if mn: return ist_to_utc(datetime(yr, mn, int(m.group(1)), 12, 0))

    m = re.match(r'^(\\d{1,2})/([A-Za-z]+)/(\\d{2,4})\\s+(\\d{1,2}):(\\d{2})\\s*(AM|PM)$', s, re.IGNORECASE)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        h, mi = int(m.group(4)), int(m.group(5)); ap = m.group(6).upper()
        if ap == "PM" and h < 12: h += 12
        if ap == "AM" and h == 12: h = 0
        mn = months.get(m.group(2).lower()[:3]); yr = int(m.group(3))
        if yr < 100: yr += 2000
        if mn: return ist_to_utc(datetime(yr, mn, int(m.group(1)), h, mi))

    m = re.match(r'^(\\d{1,2})\\s+([A-Za-z]+)\\s+(\\d{2,4}),?\\s+(\\d{1,2}):(\\d{2})\\s*(AM|PM)$', s, re.IGNORECASE)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        h, mi = int(m.group(4)), int(m.group(5)); ap = m.group(6).upper()
        if ap == "PM" and h < 12: h += 12
        if ap == "AM" and h == 12: h = 0
        mn = months.get(m.group(2).lower()[:3]); yr = int(m.group(3))
        if yr < 100: yr += 2000
        if mn: return datetime(yr, mn, int(m.group(1)), h, mi)

    m = re.match(r'^([A-Za-z]+)\\s+(\\d{1,2}),\\s+(\\d{4}),?\\s+(\\d{1,2}):(\\d{2})\\s*(AM|PM)$', s, re.IGNORECASE)
    if m:
        months = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,"jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}
        mn = months.get(m.group(1).lower()[:3]); h, mi = int(m.group(4)), int(m.group(5))
        ap = m.group(6).upper()
        if ap == "PM" and h < 12: h += 12
        if ap == "AM" and h == 12: h = 0
        if mn: return datetime(int(m.group(3)), mn, int(m.group(2)), h, mi)

    m = re.match(r'^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\\s+(\\d{1,2}):(\\d{2})\\s*(AM|PM)$', s, re.IGNORECASE)
    if m:
        days = ["monday","tuesday","wednesday","thursday","friday","saturday","sunday"]
        target = days.index(m.group(1).lower())
        h, mi = int(m.group(2)), int(m.group(3)); ap = m.group(4).upper()
        if ap == "PM" and h < 12: h += 12
        if ap == "AM" and h == 12: h = 0
        base = datetime(now_ist.year, now_ist.month, now_ist.day)
        diff = (now_ist.weekday() - target) % 7
        return ist_to_utc((base - timedelta(days=diff)).replace(hour=h, minute=mi))

    try:
        if len(s) > 5 and s[-5] in ('+','-') and s[-4:].isdigit(): s = s[:-2]+':'+s[-2:]
        s = s.replace('Z', '+00:00')
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is not None:
            from datetime import timezone
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception as e:
        print(f"[SYNC] Date parse failed '{date_str}': {e}")
        return None


def _needs_detail_fetch(existing, jira_item: dict, full_sync: bool) -> tuple[bool, str]:
    if full_sync: return True, "full_sync"
    if existing is None: return True, "new_ticket"
    if not existing.detailFetchedAt: return True, "never_fetched"
    if (existing.jiraStatus or "").strip().lower() != (jira_item.get("status") or "").strip().lower():
        return True, "status_changed"
    jira_upd = _parse_jira_date(jira_item.get("updatedAt") or "")
    if jira_upd and existing.jiraUpdatedAt and jira_upd > existing.jiraUpdatedAt:
        return True, "jira_updated_newer"
    return False, "no_changes"


def _build_final_comments(raw_jira: list, portal_comments: list) -> list[dict]:
    jira = []
    for c in raw_jira:
        author = (c.get("author") or "Jira User").strip()
        msg = (c.get("message") or "").strip()
        if not msg or author.lower() in ("automatic response", "system"): continue
        jira.append({
            "author": author, "message": msg,
            "createdAt": _parse_jira_date(c.get("createdAt") or "") or datetime.utcnow(),
            "images": c.get("images", []), "source": "jira",
        })
    portal = [
        {"id": c.id, "author": c.author, "message": c.message,
         "createdAt": c.createdAt, "images": c.images or [], "source": "portal"}
        for c in portal_comments if c.source == "portal"
    ]
    return jira + portal


async def pull_jira_updates(full_sync: bool = False) -> dict:
    service = get_jira_service()
    mode = "FULL" if full_sync else "INCREMENTAL"
    now = datetime.utcnow()
    print(f"[SYNC] {mode} sync at {now:%H:%M:%S UTC}")

    jira_tickets = service.scrape_all_tickets()
    total = len(jira_tickets)
    print(f"[SYNC] Got {total} tickets from Jira")

    all_refs = [jt.get("reference", "").strip() for jt in jira_tickets if jt.get("reference")]
    async with SessionLocal() as session:
        rows = (await session.execute(
            select(Ticket).options(selectinload(Ticket.comments)).where(Ticket.reference.in_(all_refs))
        )).scalars().all()
    existing_map = {t.reference: t for t in rows}
    print(f"[SYNC] Found {len(existing_map)} existing in PostgreSQL")

    needs_detail, skip_list = [], []
    for idx, jt in enumerate(jira_tickets):
        ref = jt.get("reference", "").strip()
        if not ref: continue
        fetch, reason = _needs_detail_fetch(existing_map.get(ref), jt, full_sync)
        (needs_detail if fetch else skip_list).append((idx, jt, existing_map.get(ref), reason))

    print(f"[SYNC] Detail fetch: {len(needs_detail)} | Skip: {len(skip_list)}")

    detail_map = {}
    if needs_detail:
        detail_map = await asyncio.to_thread(service.fetch_all_details_sync, needs_detail)
        if detail_map is None:
            raise RuntimeError("fetch_all_details_sync returned None -- browser thread crashed.")

    updated = created = skipped = 0
    errors = []
    now = datetime.utcnow()

    for idx, jt in enumerate(jira_tickets):
        ref = jt.get("reference", "").strip()
        if not ref: skipped += 1; continue
        jira_status = jt.get("status", "").strip()
        summary     = jt.get("summary", "").strip()
        created_at  = _parse_jira_date(jt.get("createdAt"))
        updated_at  = _parse_jira_date(jt.get("updatedAt"))
        list_rt     = jt.get("requestType") or ""
        existing    = existing_map.get(ref)
        vessel_name = await detect_vessel_from_text(summary) or await detect_vessel_from_text(list_rt)
        clean_summary = re.sub(
            rf"^{re.escape(vessel_name or '')}\\s*[-:]\\s*", "", summary, flags=re.IGNORECASE
        ).strip() if vessel_name else summary
        detail = detail_map.get(ref)
        new_comments_data = None
        if detail is not None:
            portal = [c for c in (existing.comments if existing else []) if c.source == "portal"]
            new_comments_data = _build_final_comments(detail.get("comments", []), portal)
        try:
            async with SessionLocal() as session:
                if existing:
                    t = (await session.execute(
                        select(Ticket).options(selectinload(Ticket.comments)).where(Ticket.reference == ref)
                    )).scalar_one()
                    t.jiraStatus = jira_status; t.jiraUrl = jt.get("url", ""); t.jiraSortOrder = idx
                    t.lastSyncedAt = now; t.updatedAt = updated_at or now; t.jiraSubmissionStatus = "SYNCED"
                    if created_at: t.jiraCreatedAt = created_at
                    if updated_at: t.jiraUpdatedAt = updated_at
                    if vessel_name: t.vesselName = vessel_name
                    if list_rt: t.requestType = list_rt
                    if detail:
                        t.detailFetchedAt = now
                        if detail.get("description"): t.description = detail["description"]
                        if detail.get("requestType"): t.requestType = detail["requestType"]
                        if detail.get("module"):      t.module = detail["module"]
                        if detail.get("environment"): t.environment = detail["environment"]
                        if detail.get("sharedWith"):  t.sharedWith = detail["sharedWith"]
                        if detail.get("attachments"): t.attachments = detail["attachments"]
                        if detail.get("raisedAt"):
                            rd = _parse_jira_date(detail["raisedAt"])
                            if rd: t.jiraCreatedAt = rd; print(f"[SYNC] raisedAt -> {rd}")
                        if new_comments_data is not None:
                            await session.execute(delete(Comment).where(Comment.ticket_id == t.id, Comment.source == "jira"))
                            for cd in new_comments_data:
                                if cd["source"] == "jira":
                                    session.add(Comment(
                                        id=str(uuid.uuid4()), ticket_id=t.id,
                                        author=cd["author"], message=cd["message"], source="jira",
                                        createdAt=cd["createdAt"], images=cd.get("images", []),
                                    ))
                    await session.commit(); updated += 1
                else:
                    t = Ticket(
                        id=str(uuid.uuid4()), reference=ref,
                        referenceNum=extract_reference_num(ref), summary=clean_summary,
                        description=(detail or {}).get("description", ""),
                        module=(detail or {}).get("module") or "Admin",
                        environment=(detail or {}).get("environment") or "Vessel",
                        priority="Minor", status="SUP IN PROGRESS", vesselName=vessel_name,
                        requester=jt.get("requester") or "", jiraStatus=jira_status,
                        jiraUrl=jt.get("url", ""), jiraSortOrder=idx, jiraSubmissionStatus="SYNCED",
                        requestType=list_rt or None,
                        attachments=(detail or {}).get("attachments", []),
                        sharedWith=(detail or {}).get("sharedWith", []),
                        createdAt=created_at or now, updatedAt=updated_at or now,
                        jiraCreatedAt=created_at, jiraUpdatedAt=updated_at,
                        lastSyncedAt=now, detailFetchedAt=now if detail else None,
                    )
                    if detail and detail.get("raisedAt"):
                        rd = _parse_jira_date(detail["raisedAt"])
                        if rd: t.jiraCreatedAt = rd
                    session.add(t); await session.flush()
                    for cd in (new_comments_data or []):
                        session.add(Comment(
                            id=str(uuid.uuid4()), ticket_id=t.id,
                            author=cd["author"], message=cd["message"], source=cd["source"],
                            createdAt=cd["createdAt"], images=cd.get("images", []),
                        ))
                    await session.commit(); created += 1
        except Exception as e:
            errors.append(f"{ref}: {e}"); print(f"[SYNC] Error {ref}: {e}")

    print(f"[SYNC] DONE -- Updated={updated} Created={created} Errors={len(errors)}")
    return {
        "mode": mode, "totalScraped": total, "detailFetched": len(detail_map),
        "detailSkipped": len(skip_list), "updated": updated, "created": created,
        "skipped": skipped, "errors": errors,
    }
""")

# ─── automation/push_service.py ──────────────────────────────────────────────
print("Step 10: Writing automation/push_service.py")
write("automation/push_service.py", """\
# automation/push_service.py
from sqlalchemy import select
from db.database import SessionLocal
from models.schema import Ticket
from automation.playwright_service import get_jira_service
from automation.status_map import extract_reference_num
from datetime import datetime


async def push_pending_tickets() -> dict:
    service = get_jira_service()

    async with SessionLocal() as session:
        pending = (await session.execute(
            select(Ticket).where(
                Ticket.jiraSubmissionStatus.in_(["PENDING", "FAILED"]),
                Ticket.reference.is_(None),
            ).limit(50)
        )).scalars().all()

    print(f"[PUSH] Found {len(pending)} tickets to submit")
    succeeded = failed = 0
    errors = []

    for ticket in pending:
        print(f"[PUSH] Processing: {ticket.summary[:50]}")
        try:
            result = service.submit_ticket({
                "id": ticket.id, "summary": ticket.summary,
                "description": ticket.description, "module": ticket.module,
                "environment": ticket.environment, "priority": ticket.priority,
                "vesselName": ticket.vesselName, "requester": ticket.requester,
            })
            reference = result.get("reference")
            jira_url  = result.get("jiraUrl")
            async with SessionLocal() as session:
                t = (await session.execute(select(Ticket).where(Ticket.id == ticket.id))).scalar_one()
                t.jiraSubmissionStatus = "SUBMITTED" if not reference else "SYNCED"
                t.jiraUrl = jira_url
                t.updatedAt = datetime.utcnow()
                if reference:
                    t.reference = reference
                    t.referenceNum = extract_reference_num(reference)
                await session.commit()
            print(f"[PUSH] SUCCESS: {reference or 'submitted'}")
            succeeded += 1
        except Exception as e:
            print(f"[PUSH] FAILED: {e}")
            errors.append(f"{ticket.summary[:40]}: {e}")
            async with SessionLocal() as session:
                t = (await session.execute(select(Ticket).where(Ticket.id == ticket.id))).scalar_one_or_none()
                if t:
                    t.jiraSubmissionStatus = "FAILED"
                    t.updatedAt = datetime.utcnow()
                    await session.commit()
            failed += 1

    return {"pushed": succeeded, "failed": failed, "errors": errors}
""")

# ─── automation/status_map.py ────────────────────────────────────────────────
print("Step 11: Writing automation/status_map.py")
write("automation/status_map.py", """\
# automation/status_map.py
MODULE_MAP = {
    "Accounts":"Accounts","Admin":"Admin","Certification":"Certification",
    "Chartering":"Chartering","Crewing":"Crewing","Dashboard":"Dashboard",
    "Data Library":"Data Library","Financial Reporting":"Financial Reporting",
    "LPSQ/HSEQ":"LPSQ/HSEQ","LiveFleet":"LiveFleet","MDM":"MDM",
    "New Applicant":"New Applicant","PMS / Maintenance":"PMS / Maintenance",
    "Payroll":"Payroll","Purchase":"Purchase","QDMS":"QDMS",
    "Replication":"Replication","Sea Roster":"Sea Roster",
    "Ticketing":"Ticketing","Training":"Training","Voyage":"Voyage",
    "SmartOps":"Admin","Maintenance":"PMS / Maintenance","Crew":"Crewing",
    "Safety":"LPSQ/HSEQ","Navigation":"LiveFleet","Inventory":"Purchase",
    "HSQE":"LPSQ/HSEQ","Dry Dock":"PMS / Maintenance","Other":"Admin",
}
ENVIRONMENT_MAP = {
    "Office":"Office","Vessel":"Vessel","Both":"Both",
    "Production":"Vessel","Staging":"Office","UAT":"Office","Development":"Office",
}
PRIORITY_TO_REQUEST_TYPE = {"Critical":"1889","Major":"1889","Minor":"1890"}
CLOSED_STATUSES = {"Cancelled","Closed","Resolved"}
KNOWN_VESSELS_FALLBACK = [
    "GCL GANGA","GCL YAMUNA","GCL SARASWATI","GCL SABARMATI",
    "GCL NARMADA","GCL TAPI","GCL FOS",
    "AM KIRTI","AM TARANG","AM UMANG",
    "AMNS POLAR","AMNS TUFMAX","AMNS MAXIMUS","AMNS STALLION",
]


def map_module(m): return MODULE_MAP.get(m, m)
def map_environment(e): return ENVIRONMENT_MAP.get(e, "Vessel")
def get_request_type_id(p): return PRIORITY_TO_REQUEST_TYPE.get(p, "1890")
def build_jira_summary(vessel, summary):
    return summary if summary.upper().startswith(vessel.upper()) else f"{vessel} - {summary}"


async def detect_vessel_from_text(text: str) -> str | None:
    if not text: return None
    vessel_names = []
    try:
        from db.database import SessionLocal
        from models.schema import Vessel
        from sqlalchemy import select
        async with SessionLocal() as session:
            rows = (await session.execute(select(Vessel.name).where(Vessel.isActive == True))).fetchall()
            vessel_names = [r[0] for r in rows if r[0]]
    except Exception as e:
        print(f"[VesselDetect] DB failed, using fallback: {e}")
        vessel_names = KNOWN_VESSELS_FALLBACK
    if not vessel_names:
        vessel_names = KNOWN_VESSELS_FALLBACK
    upper = text.upper()
    for v in sorted(vessel_names, key=len, reverse=True):
        if v.upper() in upper: return v
    return None


def extract_reference_num(reference: str | None) -> int | None:
    if not reference: return None
    import re
    m = re.search(r"-(\\d+)$", reference)
    return int(m.group(1)) if m else None
""")

# ─── migrate_mongo_to_pg.py ──────────────────────────────────────────────────
print("Step 12: Writing migrate_mongo_to_pg.py")
write("migrate_mongo_to_pg.py", """\
\"\"\"
migrate_mongo_to_pg.py
Run ONCE after first uvicorn startup to copy MongoDB data to PostgreSQL:
    python migrate_mongo_to_pg.py
\"\"\"
import asyncio, uuid, sys, os
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))

MONGO_URI = "mongodb://localhost:27017"
MONGO_DB  = "ozellar"

from db.database import SessionLocal, init_models
from models.schema import Ticket, Comment, Vessel, User


async def migrate():
    from motor.motor_asyncio import AsyncIOMotorClient
    mongo = AsyncIOMotorClient(MONGO_URI)
    mdb = mongo[MONGO_DB]
    await init_models()

    print("--- Vessels ---")
    vessels = await mdb["vessels"].find().to_list(500)
    async with SessionLocal() as s:
        for v in vessels:
            s.add(Vessel(id=str(uuid.uuid4()), name=v.get("name",""), code=v.get("code",""),
                isActive=v.get("isActive", True),
                createdAt=v.get("createdAt") or datetime.utcnow(),
                updatedAt=v.get("updatedAt") or datetime.utcnow()))
        await s.commit()
    print(f"  {len(vessels)} vessels done")

    print("--- Users ---")
    users = await mdb["users"].find().to_list(500)
    async with SessionLocal() as s:
        for u in users:
            s.add(User(id=str(uuid.uuid4()), name=u.get("name",""), email=u.get("email",""),
                password=u.get("password",""), role=u.get("role","vessel"),
                vesselName=u.get("vesselName"),
                createdAt=u.get("createdAt") or datetime.utcnow(),
                updatedAt=u.get("updatedAt") or datetime.utcnow()))
        await s.commit()
    print(f"  {len(users)} users done")

    print("--- Tickets + Comments ---")
    tickets = await mdb["tickets"].find().to_list(2000)
    tc = cc = 0
    def dt(v): return v if isinstance(v, datetime) else None
    for t in tickets:
        tid = str(uuid.uuid4())
        async with SessionLocal() as s:
            ticket = Ticket(
                id=tid, reference=t.get("reference"), referenceNum=t.get("referenceNum"),
                summary=t.get("summary",""), description=t.get("description",""),
                module=t.get("module"), environment=t.get("environment"),
                priority=t.get("priority"), requestType=t.get("requestType"),
                status=t.get("status","SUP IN PROGRESS"), jiraStatus=t.get("jiraStatus"),
                jiraSubmissionStatus=t.get("jiraSubmissionStatus","PENDING"),
                jiraUrl=t.get("jiraUrl"), jiraSortOrder=t.get("jiraSortOrder"),
                vesselName=t.get("vesselName"), requester=t.get("requester",""),
                attachments=t.get("attachments") or [], sharedWith=t.get("sharedWith") or [],
                createdAt=dt(t.get("createdAt")) or datetime.utcnow(),
                updatedAt=dt(t.get("updatedAt")) or datetime.utcnow(),
                jiraCreatedAt=dt(t.get("jiraCreatedAt")), jiraUpdatedAt=dt(t.get("jiraUpdatedAt")),
                lastSyncedAt=dt(t.get("lastSyncedAt")), detailFetchedAt=dt(t.get("detailFetchedAt")),
            )
            s.add(ticket)
            await s.flush()
            for c in (t.get("comments") or []):
                s.add(Comment(id=str(uuid.uuid4()), ticket_id=tid,
                    author=c.get("author",""), message=c.get("message",""),
                    source=c.get("source","jira"),
                    createdAt=dt(c.get("createdAt")) or datetime.utcnow(),
                    images=c.get("images") or []))
                cc += 1
            await s.commit()
            tc += 1
        if tc % 50 == 0: print(f"  ...{tc}/{len(tickets)}")

    mongo.close()
    print(f"Done! Vessels={len(vessels)} Users={len(users)} Tickets={tc} Comments={cc}")


if __name__ == "__main__":
    asyncio.run(migrate())
""")

# ─── pip install ─────────────────────────────────────────────────────────────
print()
print("Step 13: Installing Python packages...")
subprocess.run([sys.executable, "-m", "pip", "install", "sqlalchemy==2.0.30", "asyncpg==0.29.0"], check=True)

# ─── Create PostgreSQL database ───────────────────────────────────────────────
print()
print("Step 14: Creating PostgreSQL database 'ozellar'...")
try:
    import os
    env = os.environ.copy()
    env["PGPASSWORD"] = "Admin@123"
    result = subprocess.run(
        ["psql", "-U", "Deepa", "-h", "localhost", "-c", "CREATE DATABASE ozellar;"],
        env=env, capture_output=True, text=True
    )
    if "already exists" in result.stderr:
        print("  Database already exists (ok)")
    elif "CREATE DATABASE" in result.stdout:
        print("  Database created successfully")
    else:
        print("  psql output:", result.stdout or result.stderr)
except FileNotFoundError:
    print("  psql not found in PATH -- create DB manually:")
    print('  psql -U Deepa -c "CREATE DATABASE ozellar;"')

# ─── Done ─────────────────────────────────────────────────────────────────────
print()
print("=" * 55)
print("  All files written successfully!")
print("=" * 55)
print()
print("Next steps:")
print("  1. Edit .env  -- set real JWT_SECRET, JIRA_EMAIL, JIRA_PASSWORD")
print("  2. uvicorn main:app --reload    (tables auto-created on startup)")
print("  3. python migrate_mongo_to_pg.py   (copy MongoDB data - run once)")
print("  4. Open http://localhost:8000/docs")
print()
